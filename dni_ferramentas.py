"""
DNI - Ferramentas (v1)
=======================
Autor: Augusto G. Silvestrin (silvestrin) - Estagiario Eng. Produto, DNI - Metalfrio Solutions

EN: Unified DNI toolkit with 3 tabs:
  - "Projetos Preliminares": scans C:\\Projetos, generates
    "Projetos Preliminares - YYMMDD.xlsx" with KPI dashboard
  - "Extrair Itens DI": scans C:\\Projetos, extracts DI/COMP items from
    "RELACAO DE COMPONENTES" tables
  - "Cruzar PPAP": cross-references "Itens DI (Unicos)" against a Monday
    "Qualidade" export to flag PPAP availability

Requisitos: pip install openpyxl ttkbootstrap
Build: pyinstaller --onefile --windowed --name DNIFerramentas --version-file version_info.txt --icon icon_dni.ico dni_ferramentas.py
"""

__author__ = "Augusto G. Silvestrin (silvestrin)"


import os
import sys
import re
import zipfile
import datetime
import multiprocessing as mp
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import PieChart, BarChart, Reference

# =====================================================================
# CONFIGURACAO
# =====================================================================
PASTA_ORIGEM = r"C:\Projetos"
PASTA_DESTINO = str(Path.home() / "Downloads")
EXTENSOES_VALIDAS = {".xls", ".xlsx", ".xlsm"}

# Mapeamento: (aba, celula) para cada coluna de saida A..AK
CAMPOS = [
    ("Controle de Projeto",     "Q15"),  # A - PROJETO
    ("Controle de Projeto",     "Q19"),  # B - DIAS
    ("Controle de Projeto",     "Q17"),  # C - DEP
    ("Planejamento de Projeto", "E11"),  # D - STATUS%
    ("Controle de Projeto",     "Q11"),  # E - LOGIN
    ("Controle de Projeto",     "Q13"),  # F - SECUNDARIO
    ("Controle de Projeto",     "J19"),  # G - CLASSIFICACAO
    ("Controle de Projeto",     "B17"),  # H - LIDER TECNICO
    ("Controle de Projeto",     "D15"),  # I - CATEGORIA
    ("Controle de Projeto",     "G13"),  # J - PROGRAMA
    ("Controle de Projeto",     "D17"),  # K - FASE PROJETO
    ("Controle de Projeto",     "J15"),  # L - DEP.SOLICITANTE
    ("Controle de Projeto",     "B11"),  # M - ASSUNTO
    ("Controle de Projeto",     "B15"),  # N - MODELO
    ("Controle de Projeto",     "J17"),  # O - SOLICITACAO
    ("Controle de Projeto",     "G17"),  # P - ESTIMATIVA (formula)
    ("Planejamento de Projeto", "G20"),  # Q - ESTIMATIVA.DNI
]

# Cabecalhos fixos (parte A..Q) e finais (REV.01-10). COD.PROD/PR ficam
# entre eles, mas seu NUMERO e' dinamico (depende do maximo encontrado
# entre todos os arquivos) -> CABECALHOS completo e' montado em main().
CABECALHOS_FIXOS = [
    "PROJETO", "DIAS", "DEP", "STATUS%", "LOGIN", "SECUNDÁRIO",
    "CLASSIFICAÇÃO", "LÍDER TÉCNICO", "CATEGORIA", "PROGRAMA",
    "FASE PROJETO", "DEP.SOLICITANTE", "ASSUNTO", "MODELO",
    "SOLICITAÇÃO", "ESTIMATIVA", "ESTIMATIVA.DNI",
]
CABECALHOS_REV = [
    "REV.00", "REV.01", "REV.02", "REV.03", "REV.04", "REV.05",
    "REV.06", "REV.07", "REV.08", "REV.09", "REV.10",
]

N_CAMPOS_FIXOS = len(CAMPOS)          # 17 (A..Q)
N_REV = len(CABECALHOS_REV)           # 10

# Texto procurado para localizar a tabela de produtos acabados na aba
# "Controle de Projeto" (busca dinamica, pois a posicao varia por arquivo)
TEXTO_TABELA_PRODUTOS = "CÓDIGO PRODUTO ACABADO"




def extrair_revisoes(caminho):
    """
    EN: Dynamically locates the "Revisões de Projeto/DEP" table inside the
    "Revisões DEP" sheet (header row with REVISÃO/EFETIVAÇÃO/EMISSÃO/
    LIBERADO columns), then reads the EMISSÃO (date) column for revisions
    01-10 (skipping row "00").

    Returns a list of length N_REV (10), padded with None.
    """
    revisoes = [None] * N_REV

    try:
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    except Exception:
        return revisoes

    try:
        if "Revisões DEP" not in wb.sheetnames:
            return revisoes
        ws = wb["Revisões DEP"]

        linha_header = None
        col_revisao = None
        col_emissao = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100)), start=1):
            achou_revisao = False
            achou_emissao = False
            for col_idx, cell in enumerate(row, start=1):
                if not isinstance(cell.value, str):
                    continue
                texto = cell.value.strip().upper()
                if texto == "REVISÃO":
                    col_revisao = col_idx
                    achou_revisao = True
                elif texto == "EMISSÃO":
                    col_emissao = col_idx
                    achou_emissao = True
            if achou_revisao and achou_emissao:
                linha_header = row_idx
                break

        if linha_header is None or col_revisao is None or col_emissao is None:
            return revisoes

        # Linhas de dados: REVISAO "00" a "10" -> REV.00..REV.10 (11 colunas)
        for linha in range(linha_header + 1, min(linha_header + 13, ws.max_row + 1)):
            rev_label = ws.cell(row=linha, column=col_revisao).value
            if rev_label is None:
                continue
            rev_str = str(rev_label).strip()
            try:
                idx = int(rev_str)  # "00" -> indice 0, "01" -> indice 1, ...
            except ValueError:
                continue
            if 0 <= idx < N_REV:
                revisoes[idx] = ws.cell(row=linha, column=col_emissao).value
    finally:
        wb.close()

    return revisoes


def _norm(texto):
    """EN: Normalize text for robust matching: uppercase, strip accents,
    collapse internal whitespace."""
    if not isinstance(texto, str):
        return ""
    s = texto.strip().upper()
    substituicoes = {
        "Á": "A", "À": "A", "Â": "A", "Ã": "A",
        "É": "E", "È": "E", "Ê": "E",
        "Í": "I", "Ì": "I",
        "Ó": "O", "Ò": "O", "Ô": "O", "Õ": "O",
        "Ú": "U", "Ù": "U",
        "Ç": "C",
    }
    for k, v in substituicoes.items():
        s = s.replace(k, v)
    return " ".join(s.split())  # colapsa espacos multiplos


def extrair_produtos_acabados(caminho):
    """
    EN: Dynamically locates the "CÓDIGO PRODUTO ACABADO" table inside the
    "Controle de Projeto" sheet (its row position AND length vary between
    files), then reads ALL rows from the "CÓDIGO" and "FICHA TÉCNICA" (PR)
    columns of that table (stops at first empty CÓDIGO cell).

    Returns (cod_prods, prs, diag) — variable-length lists (same length),
    NOT padded, plus a short diagnostic string ("ok", "tabela nao
    encontrada", "cabecalho nao encontrado", "sem aba Controle de
    Projeto", "erro ao abrir"), used by main() to report files where
    nothing was extracted.
    """
    cod_prods = []
    prs = []

    try:
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    except Exception:
        return cod_prods, prs, "erro ao abrir"

    try:
        if "Controle de Projeto" not in wb.sheetnames:
            return cod_prods, prs, "sem aba 'Controle de Projeto'"
        ws = wb["Controle de Projeto"]

        TEXTO_TABELA_NORM = _norm(TEXTO_TABELA_PRODUTOS)

        linha_titulo = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)), start=1):
            for cell in row:
                if TEXTO_TABELA_NORM in _norm(cell.value):
                    linha_titulo = row_idx
                    break
            if linha_titulo:
                break

        if linha_titulo is None:
            return cod_prods, prs, "titulo 'CODIGO PRODUTO ACABADO' nao encontrado"

        # Cabecalho da tabela costuma ficar 1 linha abaixo do titulo, mas
        # busca numa janela de algumas linhas para tolerar variacoes
        # (linhas em branco extras, mesclas, etc.)
        linha_header = None
        col_codigo = None
        col_ficha = None
        for candidata in range(linha_titulo + 1, min(linha_titulo + 6, ws.max_row + 1)):
            cc = None
            cf = None
            cd = None  # col_denominacao
            for col_idx, cell in enumerate(next(ws.iter_rows(min_row=candidata, max_row=candidata)), start=1):
                texto = _norm(cell.value)
                if texto == "CODIGO" and cc is None:
                    cc = col_idx
                elif "FICHA" in texto and "TECNICA" in texto:
                    cf = col_idx
                elif texto == "DENOMINACAO" and cd is None:
                    cd = col_idx
            if cc is not None:
                linha_header = candidata
                col_codigo = cc
                # FICHA TECNICA pode estar vazia no cabecalho (botao/forma
                # sobreposto, sem texto na celula) -> usa posicao relativa
                # a DENOMINACAO (sempre a coluna seguinte), se nao achou pelo texto
                col_ficha = cf if cf is not None else (cd + 1 if cd is not None else None)
                break

        if linha_header is None or col_codigo is None:
            return cod_prods, prs, f"cabecalho 'CODIGO' nao encontrado (titulo na linha {linha_titulo})"

        if col_ficha is None and linha_header > 1:
            # Pode estar mesclada com a linha acima do header
            for col_idx, cell in enumerate(next(ws.iter_rows(min_row=linha_header - 1, max_row=linha_header - 1)), start=1):
                texto = _norm(cell.value)
                if "FICHA" in texto and "TECNICA" in texto:
                    col_ficha = col_idx
                    break

        diag_ficha_dump = None
        if col_ficha is None:
            dump = []
            for r in range(max(1, linha_header - 1), min(linha_header + 2, ws.max_row + 1)):
                for col_idx, cell in enumerate(next(ws.iter_rows(min_row=r, max_row=r)), start=1):
                    if cell.value is not None and str(cell.value).strip() != "":
                        dump.append(f"L{r}C{col_idx}={cell.value!r}")
            diag_ficha_dump = "; ".join(dump[:30])

        # Le todas as linhas de dados ate encontrar CODIGO vazio
        MAX_ITENS_SEGURANCA = 100  # limite de seguranca para evitar loop infinito
        for i in range(MAX_ITENS_SEGURANCA):
            linha_dado = linha_header + 1 + i
            if linha_dado > ws.max_row:
                break
            valor_codigo = ws.cell(row=linha_dado, column=col_codigo).value
            if valor_codigo is None or str(valor_codigo).strip() == "":
                break
            cod_prods.append(valor_codigo)
            prs.append(ws.cell(row=linha_dado, column=col_ficha).value if col_ficha else None)

        if not cod_prods:
            diag = (f"tabela encontrada (titulo linha {linha_titulo}, header linha "
                    f"{linha_header}, col CODIGO={col_codigo}, col FICHA={col_ficha}) "
                    f"mas 0 codigos lidos (primeira celula CODIGO vazia)")
        elif col_ficha is None:
            diag = (f"COD.PROD ok ({len(cod_prods)} itens), mas coluna 'FICHA TECNICA' "
                    f"NAO encontrada no header (linha {linha_header}) -> PRs ficam vazios. "
                    f"Dump das celulas proximas: {diag_ficha_dump}")
        else:
            diag = "ok"
    except Exception as e:
        return cod_prods, prs, f"erro: {e}"
    finally:
        wb.close()

    return cod_prods, prs, diag


COL_DEP = 3
COL_STATUS_PCT = 4
COL_ASSUNTO = 13  # M — alinhamento especial: esquerda + meio

# Cores identidade Metalfrio (cabecalho de todas as tabelas/abas)
AZUL_METALFRIO = "002F6C"   # azul marinho logo
VERDE_METALFRIO = "00A651"  # verde esmeralda logo
COL_SOLICITACAO = 15
COL_ESTIMATIVA = 16
COL_ESTIMATIVA_DNI = 17

LINHA_HEADER = 1
LINHA_PRIMEIRA_DADO = 2

STATUS_TEXTUAIS = ["PRELIMINAR", "CONCEITO", "CONCEITO/PRÉ ALPHA", "PRÉ ALPHA", "CANCELADO"]


def listar_arquivos(pasta):
    arquivos = []
    for entry in os.scandir(pasta):
        if entry.is_file():
            ext = Path(entry.name).suffix.lower()
            if ext in EXTENSOES_VALIDAS and not entry.name.startswith("~$"):
                arquivos.append(entry.path)
    return arquivos


def ler_arquivo(caminho):
    """
    EN: Worker function — reads fixed cells (CAMPOS, 17 fields A..Q) via
    fast XML parsing (fallback openpyxl), then dynamically locates the
    "CÓDIGO PRODUTO ACABADO" table (variable-length COD.PROD/PR) and the
    "Revisões DEP" table (REV.01-10) via openpyxl.
    Must be a top-level function (picklable) for multiprocessing.
    Returns (caminho, dados_dict_ou_None, erro_str_ou_None), where
    dados_dict = {"fixos": [...17], "cod_prods": [...N], "prs": [...N],
                   "revisoes": [...10]}.
    """
    valores_fixos = None
    try:
        valores_fixos = ler_arquivo_xml_rapido(caminho)
    except Exception:
        pass

    if valores_fixos is None:
        valores_fixos = [None] * N_CAMPOS_FIXOS
        try:
            wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
        except Exception as e:
            return (caminho, None, str(e))

        try:
            for idx, (aba, celula) in enumerate(CAMPOS):
                if aba in wb.sheetnames:
                    try:
                        valores_fixos[idx] = wb[aba][celula].value
                    except Exception:
                        valores_fixos[idx] = None
        finally:
            wb.close()

    cod_prods, prs, diag_produtos = extrair_produtos_acabados(caminho)
    revisoes = extrair_revisoes(caminho)

    dados = {
        "fixos": valores_fixos,
        "cod_prods": cod_prods,
        "prs": prs,
        "revisoes": revisoes,
        "diag_produtos": diag_produtos,
    }
    return (caminho, dados, None)


# ---- Regex pre-compilados (reuso entre chamadas) ----
_RE_CELL_REF = re.compile(r'([A-Z]+)(\d+)')
_RE_SHEET_TAG = re.compile(r'<sheet\b[^>]*name="([^"]*)"[^>]*r:id="([^"]*)"', re.IGNORECASE)
_RE_SHEET_TAG_ALT = re.compile(r'<sheet\b[^>]*r:id="([^"]*)"[^>]*name="([^"]*)"', re.IGNORECASE)
_RE_RELATIONSHIP = re.compile(r'<Relationship\b[^>]*/?>')
_RE_ATTR = re.compile(r'(\w+)="([^"]*)"')
_RE_SHARED_SI = re.compile(r'<si\b[^>]*>(.*?)</si>', re.DOTALL)
_RE_T_TAG = re.compile(r'<t\b[^>]*>(.*?)</t>', re.DOTALL)


def _split_cell_ref(ref):
    m = _RE_CELL_REF.match(ref)
    if not m:
        return None, None
    col_letters, row_str = m.groups()
    return column_index_from_string(col_letters), int(row_str)


def _unescape_xml(s):
    return (s.replace("&amp;", "&").replace("&lt;", "<")
              .replace("&gt;", ">").replace("&quot;", '"')
              .replace("&apos;", "'"))


def _parse_shared_strings(xml_bytes):
    """EN: Parses xl/sharedStrings.xml into a list of strings (index -> text)."""
    text = xml_bytes.decode("utf-8", errors="replace")
    out = []
    for m in _RE_SHARED_SI.finditer(text):
        chunk = m.group(1)
        # Concatena todos os <t> dentro do <si> (cobre runs com formatacao <r>)
        parts = _RE_T_TAG.findall(chunk)
        out.append(_unescape_xml("".join(parts)))
    return out


def _build_sheet_name_map(workbook_xml, rels_xml):
    """
    EN: Maps sheet display name (e.g. 'Controle de Projeto') to its
    internal zip path (e.g. 'xl/worksheets/sheet2.xml'), using
    workbook.xml (name + r:id) and workbook.xml.rels (r:id + Target).
    """
    wb_text = workbook_xml.decode("utf-8", errors="replace")
    name_to_rid = {}
    for m in _RE_SHEET_TAG.finditer(wb_text):
        name, rid = m.group(1), m.group(2)
        name_to_rid[_unescape_xml(name)] = rid
    if not name_to_rid:
        for m in _RE_SHEET_TAG_ALT.finditer(wb_text):
            rid, name = m.group(1), m.group(2)
            name_to_rid[_unescape_xml(name)] = rid

    rels_text = rels_xml.decode("utf-8", errors="replace")
    rid_to_target = {}
    for tag in _RE_RELATIONSHIP.findall(rels_text):
        attrs = dict(_RE_ATTR.findall(tag))
        if "Id" in attrs and "Target" in attrs:
            rid_to_target[attrs["Id"]] = attrs["Target"]

    name_to_path = {}
    for name, rid in name_to_rid.items():
        target = rid_to_target.get(rid)
        if target:
            if target.startswith("/"):
                path = target.lstrip("/")
            elif target.startswith("xl/"):
                path = target
            else:
                path = "xl/" + target
            name_to_path[name] = path
    return name_to_path


def _extract_cell_value(sheet_xml_bytes, cell_ref, shared_strings):
    """
    EN: Extracts the value of a single cell from worksheet XML bytes.
    Returns python value (str/float/int/datetime/None) without loading
    the full sheet into a DOM (regex-based, fast for sparse lookups).
    For formula cells, returns the cached <v> result (same as
    openpyxl data_only=True).
    """
    text = sheet_xml_bytes.decode("utf-8", errors="replace")

    # Localiza a tag <c r="REF" ...> ... </c> (ou self-closing)
    pattern = re.compile(
        r'<c\s+[^>]*\br="' + re.escape(cell_ref) + r'"[^>]*?(?:/>|>(.*?)</c>)',
        re.DOTALL,
    )
    m = pattern.search(text)
    if not m:
        return None

    full_tag_match = re.search(
        r'<c\s+[^>]*\br="' + re.escape(cell_ref) + r'"[^>]*?(?=/>|>)',
        text,
    )
    full_tag = full_tag_match.group(0) if full_tag_match else ""

    type_match = re.search(r'\bt="([^"]*)"', full_tag)
    cell_type = type_match.group(1) if type_match else None

    inner = m.group(1) or ""

    # Valor (cache de formula tambem fica em <v>)
    v_match = re.search(r'<v>(.*?)</v>', inner, re.DOTALL)
    raw_value = v_match.group(1) if v_match else None
    if raw_value is not None and raw_value.strip() == "":
        raw_value = None

    if raw_value is None:
        # Pode ser string inline <is><t>...</t></is> (t="inlineStr" ou "str")
        is_match = re.search(r'<is>(.*?)</is>', inner, re.DOTALL)
        if is_match:
            t_parts = _RE_T_TAG.findall(is_match.group(1))
            return _unescape_xml("".join(t_parts))
        return None

    if cell_type == "s":
        # Shared string index
        try:
            idx = int(raw_value)
            return shared_strings[idx] if 0 <= idx < len(shared_strings) else None
        except (ValueError, IndexError):
            return None
    elif cell_type == "str":
        return _unescape_xml(raw_value)
    elif cell_type == "b":
        return raw_value == "1"
    else:
        # Numerico (inclui datas serializadas como numero)
        try:
            num = float(raw_value)
        except ValueError:
            return _unescape_xml(raw_value)

        if num == int(num):
            num_int = int(num)
            # Heuristica: numeros nessa faixa em celulas de data conhecidas
            # sao convertidos para datetime (serial date do Excel, base 1899-12-30)
            return _maybe_excel_date(num_int, num)
        return num


# Colunas de CAMPOS (1-based, dentro de cada aba) que sao datas no arquivo
# individual de origem -> usado para decidir se converte serial->datetime
_CELULAS_DATA_ORIGEM = {"O15", "G17", "G20", "D14", "D15", "D16", "D17", "D18",
                         "D19", "D20", "D21", "D22", "D23", "J17"}


def _maybe_excel_date(num_int, num_float):
    """
    EN: Excel stores dates as serial numbers (days since 1899-12-30).
    Converts plausible date-range integers to datetime; otherwise returns
    the original number. Range chosen to cover ~1950-2100.
    """
    if 18000 <= num_int <= 73000:  # ~1949-04 .. ~2099-11
        try:
            base = datetime.datetime(1899, 12, 30)
            return base + datetime.timedelta(days=num_int)
        except OverflowError:
            return num_float
    return num_float if num_float != num_int else num_int


def ler_arquivo_xml_rapido(caminho):
    """
    EN: Fast path — reads CAMPOS directly from the .xlsx/.xlsm zip's XML,
    without building the openpyxl object model. Returns None if the file
    isn't a valid OOXML zip (e.g. legacy .xls), so caller can fall back.
    """
    ext = Path(caminho).suffix.lower()
    if ext == ".xls":
        return None  # formato binario antigo - usa fallback openpyxl

    with zipfile.ZipFile(caminho, "r") as z:
        names = set(z.namelist())

        if "xl/workbook.xml" not in names or "xl/_rels/workbook.xml.rels" not in names:
            return None

        workbook_xml = z.read("xl/workbook.xml")
        rels_xml = z.read("xl/_rels/workbook.xml.rels")
        name_to_path = _build_sheet_name_map(workbook_xml, rels_xml)

        shared_strings = []
        if "xl/sharedStrings.xml" in names:
            shared_strings = _parse_shared_strings(z.read("xl/sharedStrings.xml"))

        # Cache de XML de aba (cada aba pode ter varias celulas pedidas)
        sheet_xml_cache = {}
        valores = [None] * len(CAMPOS)

        for idx, (aba, celula) in enumerate(CAMPOS):
            sheet_path = name_to_path.get(aba)
            if not sheet_path or sheet_path not in names:
                valores[idx] = None
                continue

            if sheet_path not in sheet_xml_cache:
                sheet_xml_cache[sheet_path] = z.read(sheet_path)

            valores[idx] = _extract_cell_value(
                sheet_xml_cache[sheet_path], celula, shared_strings
            )

        return valores


def main(pasta_origem=None, pasta_destino=None, log=print, progress=None):
    """
    EN: log is a callable(str) used for all progress messages — defaults
    to print (console mode), but the GUI passes a function that appends
    to a text widget instead.
    progress, if given, is a callable(current, total, message) called
    as files are processed and at key stages, for driving a progress bar.
    """
    pasta_origem = pasta_origem or PASTA_ORIGEM
    pasta_destino = pasta_destino or PASTA_DESTINO

    def _progress(current, total, message=""):
        if progress:
            progress(current, total, message)

    log("=" * 70)
    log("Geracao de 'Projetos Preliminares - YYMMDD.xlsx' (v2 - paralelo)")
    log("=" * 70)

    if not os.path.isdir(pasta_origem):
        log(f"ERRO: pasta de origem nao encontrada: {pasta_origem}")
        return None

    arquivos = listar_arquivos(pasta_origem)
    total = len(arquivos)
    log(f"Encontrados {total} arquivo(s) em {pasta_origem}\n")
    _progress(0, total, "Iniciando...")

    if total == 0:
        log("Nenhum arquivo para processar. Encerrando.")
        return None

    inicio = datetime.datetime.now()

    n_workers = max(1, min(os.cpu_count() or 4, 8))
    log(f"Lendo arquivos em paralelo usando {n_workers} processo(s)...\n")

    resultados = {}
    erros = []
    with mp.Pool(processes=n_workers) as pool:
        for i, (caminho, valores, erro) in enumerate(pool.imap_unordered(ler_arquivo, arquivos), start=1):
            nome = os.path.basename(caminho)
            if erro:
                erros.append((nome, erro))
                log(f"[{i}/{total}] ERRO: {nome} -> {erro}")
            else:
                resultados[caminho] = valores
                log(f"[{i}/{total}] OK: {nome}")
                diag = valores.get("diag_produtos")
                if diag and diag != "ok":
                    log(f"    -> COD.PROD/PR: {diag}")
            _progress(i, total, f"Lendo arquivos... ({i}/{total})")

    tempo_leitura = (datetime.datetime.now() - inicio).total_seconds()
    log(f"\nLeitura concluida em {tempo_leitura:.1f}s "
        f"({len(resultados)} ok, {len(erros)} com erro)")

    # --- Descobre o numero maximo de codigos de produto entre todos os
    #     arquivos, para montar colunas COD.PROD1..N / PR1..N dinamicas ---
    max_cod_prod = 0
    for caminho in arquivos:
        dados = resultados.get(caminho)
        if dados is None:
            continue
        max_cod_prod = max(max_cod_prod, len(dados.get("cod_prods", [])))
    max_cod_prod = max(max_cod_prod, 1)  # garante pelo menos 1 coluna

    cabecalhos_cod = [f"COD.PROD{i}" for i in range(1, max_cod_prod + 1)]
    cabecalhos_pr = [f"PR{i}" for i in range(1, max_cod_prod + 1)]
    cabecalhos = CABECALHOS_FIXOS + cabecalhos_cod + cabecalhos_pr + CABECALHOS_REV
    total_colunas = len(cabecalhos)

    if max_cod_prod > 5:
        log(f"Aviso: encontrados projetos com até {max_cod_prod} códigos de produto "
              f"(colunas COD.PROD1-{max_cod_prod} / PR1-{max_cod_prod} criadas).")

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "Projetos"

    for j, titulo in enumerate(cabecalhos, start=1):
        ws.cell(row=LINHA_HEADER, column=j, value=titulo)

    linha_atual = LINHA_PRIMEIRA_DADO
    for caminho in arquivos:
        dados = resultados.get(caminho)
        if dados is None:
            continue

        cod_prods = dados["cod_prods"] + [None] * (max_cod_prod - len(dados["cod_prods"]))
        prs = dados["prs"] + [None] * (max_cod_prod - len(dados["prs"]))

        linha_valores_fixas = (list(dados["fixos"]) + list(cod_prods))
        n_cols_fixas = len(linha_valores_fixas)

        for j, v in enumerate(linha_valores_fixas, start=1):
            ws.cell(row=linha_atual, column=j, value=v)

        # Colunas PR: texto simples (ex: "PR 26-078"), sem hyperlink
        for k, texto_pr in enumerate(prs):
            ws.cell(row=linha_atual, column=n_cols_fixas + 1 + k, value=texto_pr)

        # Colunas REV
        col_rev_base = n_cols_fixas + len(prs)
        for k, v in enumerate(dados["revisoes"]):
            ws.cell(row=linha_atual, column=col_rev_base + 1 + k, value=v)

        linha_atual += 1

    ultima_linha = linha_atual - 1
    log(f"Total de linhas de dados gravadas: {ultima_linha - LINHA_PRIMEIRA_DADO + 1}")

    # --- Recalcula DIAS com a data atual (evita depender do cache da formula) ---
    recalcular_dias(ws, ultima_linha)

    # Indices (1-based) das colunas de data REV.01..N (dependem de max_cod_prod)
    col_rev_inicio = len(CABECALHOS_FIXOS) + 2 * max_cod_prod + 1
    cols_rev = list(range(col_rev_inicio, col_rev_inicio + N_REV))
    colunas_data = {COL_SOLICITACAO, COL_ESTIMATIVA, COL_ESTIMATIVA_DNI} | set(cols_rev)

    resumo = calcular_resumo(ws, ultima_linha)
    aplicar_formatacao(ws, ultima_linha, cabecalhos, total_colunas, colunas_data)
    criar_aba_efetivados(wb_out, ws, ultima_linha)
    criar_aba_preliminares(wb_out, ws, ultima_linha)
    criar_aba_resumo(wb_out, resumo)

    hoje = datetime.date.today().strftime("%y%m%d")
    nome_saida = f"Projetos Preliminares - {hoje}.xlsx"
    caminho_saida = os.path.join(pasta_destino, nome_saida)

    wb_out.save(caminho_saida)

    tempo_total = (datetime.datetime.now() - inicio).total_seconds()
    log(f"\nArquivo salvo em: {caminho_saida}")
    log(f"Tempo total: {tempo_total:.1f}s")
    if erros:
        log(f"\n{len(erros)} arquivo(s) com erro (pulados):")
        for nome, msg in erros:
            log(f"  - {nome}: {msg}")
    log("Concluido.")
    _progress(total, total, "Concluido")
    return caminho_saida


COL_CATEGORIA = 9  # I
COL_LIDER_TECNICO = 8  # H


def calcular_resumo(ws, ultima_linha):
    """
    Calcula estatisticas de resumo (sem escrever na planilha), separando
    duas populacoes:

    - EFETIVADOS: DEP (coluna C) e' uma data (projeto ja concluido).
      DIAS = ESTIMATIVA - DEP (negativo = atrasou na efetivacao).
    - EM ABERTO: DEP == "PRELIMINAR" (projeto ainda em andamento).
      DIAS = ESTIMATIVA - HOJE() (negativo = estimativa ja vencida).

    Para cada populacao: total, no prazo (qtd/%), atrasados (qtd/%),
    atraso medio (dias, media dos |DIAS| dos atrasados).

    'Atrasados por categoria': soma de (efetivados atrasados + em aberto
    atrasados), agrupados pela coluna CATEGORIA.

    Retorna dict com chaves:
      'contagens'   - lista (rotulo, valor) por STATUS_TEXTUAIS + EFETIVADOS + TOTAL
      'efetivados'  - dict com total/no_prazo/atrasados/pct_no_prazo/pct_atrasados/atraso_medio_dias
      'em_aberto'   - dict com as mesmas chaves
      'atrasados_categoria' - lista (categoria, qtd) ordenada desc
      'prazo'       - lista (rotulo, pct) para o donut [NO PRAZO, FORA DO PRAZO] (efetivados)
    """
    contagens = {s: 0 for s in STATUS_TEXTUAIS}

    efet_total = 0
    efet_no_prazo = 0
    efet_atrasos = []  # dias de atraso (positivos) dos efetivados atrasados

    aberto_total = 0
    aberto_no_prazo = 0
    aberto_atrasos = []

    atrasados_categoria = {}  # categoria -> qtd (efetivados + em aberto atrasados)
    projetos_por_lider = {}   # lider tecnico -> qtd (todos os projetos, qualquer status)

    for row in range(LINHA_PRIMEIRA_DADO, ultima_linha + 1):
        dep_val = ws.cell(row=row, column=COL_DEP).value
        dias_val = ws.cell(row=row, column=COL_DIAS).value
        categoria = ws.cell(row=row, column=COL_CATEGORIA).value
        categoria = str(categoria).strip() if categoria else "(sem categoria)"

        lider = ws.cell(row=row, column=COL_LIDER_TECNICO).value
        lider = str(lider).strip() if lider else "(sem líder)"
        projetos_por_lider[lider] = projetos_por_lider.get(lider, 0) + 1

        if dep_val is None:
            continue

        if isinstance(dep_val, (datetime.date, datetime.datetime)):
            # --- EFETIVADO ---
            efet_total += 1
            if isinstance(dias_val, (int, float)):
                if dias_val >= 0:
                    efet_no_prazo += 1
                else:
                    efet_atrasos.append(-dias_val)
                    atrasados_categoria[categoria] = atrasados_categoria.get(categoria, 0) + 1
        else:
            chave = str(dep_val).strip().upper()
            if chave in contagens:
                contagens[chave] += 1

            if chave == "PRELIMINAR":
                # --- EM ABERTO ---
                aberto_total += 1
                if isinstance(dias_val, (int, float)):
                    if dias_val >= 0:
                        aberto_no_prazo += 1
                    else:
                        aberto_atrasos.append(-dias_val)
                        atrasados_categoria[categoria] = atrasados_categoria.get(categoria, 0) + 1

    total_itens = ultima_linha - LINHA_PRIMEIRA_DADO + 1

    def _bloco(total, no_prazo, atrasos):
        atrasados = total - no_prazo
        pct_no_prazo = (no_prazo / total) if total > 0 else 0
        pct_atrasados = (atrasados / total) if total > 0 else 0
        atraso_medio = (sum(atrasos) / len(atrasos)) if atrasos else 0
        return {
            "total": total,
            "no_prazo": no_prazo,
            "atrasados": atrasados,
            "pct_no_prazo": pct_no_prazo,
            "pct_atrasados": pct_atrasados,
            "atraso_medio_dias": round(atraso_medio, 1),
        }

    bloco_efetivados = _bloco(efet_total, efet_no_prazo, efet_atrasos)
    bloco_aberto = _bloco(aberto_total, aberto_no_prazo, aberto_atrasos)

    itens_contagens = [(s, contagens[s]) for s in STATUS_TEXTUAIS]
    itens_contagens.append(("EFETIVADOS", efet_total))
    itens_contagens.append(("TOTAL", total_itens))

    itens_prazo = [
        ("NO PRAZO", bloco_efetivados["pct_no_prazo"]),
        ("FORA DO PRAZO", bloco_efetivados["pct_atrasados"]),
    ]

    itens_atrasados_categoria = sorted(atrasados_categoria.items(), key=lambda kv: -kv[1])
    itens_projetos_por_lider = sorted(projetos_por_lider.items(), key=lambda kv: -kv[1])

    return {
        "contagens": itens_contagens,
        "efetivados": bloco_efetivados,
        "em_aberto": bloco_aberto,
        "atrasados_categoria": itens_atrasados_categoria,
        "projetos_por_lider": itens_projetos_por_lider,
        "prazo": itens_prazo,
    }


def criar_aba_resumo(wb_out, resumo):
    """
    EN: Creates a "Resumo" sheet as a dashboard:
      - Top: 8 KPI "cards" (grid 4x2, merged cells, large font, colored
        fill), split into two groups:
          EFETIVADOS: total, no prazo (qtd/%), atrasados (qtd/%),
                       atraso medio (dias)
          EM ABERTO (PRELIMINAR): same 4 indicators
      - Below: data tables (STATUS, PRAZO, ATRASADOS POR CATEGORIA)
        feeding three charts: pie (status distribution), donut
        (efetivados no prazo x fora do prazo), and a horizontal bar
        chart (atrasados por categoria).
    """
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import CharacterProperties, ParagraphProperties, Paragraph

    ws = wb_out.create_sheet("Resumo", 0)  # primeira aba

    # ================= PALETA DARK (estilo "Dashboard Projetos") =================
    BG_ESCURO = "1B1F24"        # fundo geral da aba
    BG_CARD = "262B33"          # fundo dos cards/tabelas
    TEAL = "2EC4B6"             # acento principal (numeros, series principais)
    TEAL_CLARO = "5FE3D4"
    CINZA_TEXTO = "C7CCD1"      # texto secundario (labels)
    BRANCO = "F4F6F8"           # texto principal
    LARANJA = "FF8A5B"          # acento secundario (atrasos)
    AMARELO = "FFD166"
    AZUL = "5C9DFF"

    fonte_header = Font(bold=True, color=VERDE_METALFRIO)
    fill_header = PatternFill(start_color=AZUL_METALFRIO, end_color=AZUL_METALFRIO, fill_type="solid")
    fina = Side(style="thin", color="3A4048")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)

    # Preenche um range generoso com o fundo escuro (linhas/colunas usadas
    # pelo dashboard + margem), para simular o tema dark da referencia
    fill_bg = PatternFill(start_color=BG_ESCURO, end_color=BG_ESCURO, fill_type="solid")
    for r in range(1, 45):
        for c in range(1, 32):
            ws.cell(row=r, column=c).fill = fill_bg

    ef = resumo["efetivados"]
    ab = resumo["em_aberto"]

    # ================= CARDS (DASHBOARD) - grid 4x2 =================
    # Linha 1: EFETIVADOS | Linha 2: EM ABERTO (PRELIMINAR)
    # cor = cor do NUMERO (acento); fundo do card e' sempre BG_CARD
    CARDS = [
        ("EFETIVADOS (TOTAL)", ef["total"], "0", AZUL),
        ("EFETIVADOS NO PRAZO", ef["pct_no_prazo"], "0.0%", TEAL),
        ("EFETIVADOS ATRASADOS", ef["pct_atrasados"], "0.0%", LARANJA),
        ("ATRASO MÉDIO - EFETIVADOS (DIAS)", ef["atraso_medio_dias"], "0.0", LARANJA),

        ("EM ABERTO (PRELIMINAR)", ab["total"], "0", AMARELO),
        ("EM ABERTO NO PRAZO", ab["pct_no_prazo"], "0.0%", TEAL),
        ("EM ABERTO ATRASADOS", ab["pct_atrasados"], "0.0%", LARANJA),
        ("ATRASO MÉDIO - EM ABERTO (DIAS)", ab["atraso_medio_dias"], "0.0", LARANJA),
    ]

    CARD_LARGURA_COLS = 3   # cada card ocupa 3 colunas
    CARDS_POR_LINHA = 4

    for idx, (rotulo, valor, fmt, cor) in enumerate(CARDS):
        linha_base = 1 + (idx // CARDS_POR_LINHA) * 3  # 3 linhas por "linha" de cards (num + label + espaco)
        col_base = 1 + (idx % CARDS_POR_LINHA) * CARD_LARGURA_COLS

        # Linha do numero (mesclada)
        ws.merge_cells(start_row=linha_base, start_column=col_base,
                        end_row=linha_base, end_column=col_base + CARD_LARGURA_COLS - 1)
        c_num = ws.cell(row=linha_base, column=col_base, value=valor)
        c_num.number_format = fmt
        c_num.font = Font(bold=True, size=22, color=cor)
        c_num.alignment = Alignment(horizontal="center", vertical="center")
        c_num.fill = PatternFill(start_color=BG_CARD, end_color=BG_CARD, fill_type="solid")

        # Linha do rotulo (mesclada)
        ws.merge_cells(start_row=linha_base + 1, start_column=col_base,
                        end_row=linha_base + 1, end_column=col_base + CARD_LARGURA_COLS - 1)
        c_lbl = ws.cell(row=linha_base + 1, column=col_base, value=rotulo)
        c_lbl.font = Font(bold=True, size=9, color=CINZA_TEXTO)
        c_lbl.alignment = Alignment(horizontal="center", vertical="center")
        c_lbl.fill = PatternFill(start_color=BG_CARD, end_color=BG_CARD, fill_type="solid")

        # Bordas em toda a area do card
        for r in (linha_base, linha_base + 1):
            for c in range(col_base, col_base + CARD_LARGURA_COLS):
                ws.cell(row=r, column=c).border = borda

    # Define alturas das linhas dos cards (numero maior, rotulo menor)
    n_linhas_cards = ((len(CARDS) - 1) // CARDS_POR_LINHA + 1)
    for k in range(n_linhas_cards):
        ws.row_dimensions[1 + k * 3].height = 34       # linha do numero
        ws.row_dimensions[2 + k * 3].height = 16       # linha do rotulo
        ws.row_dimensions[3 + k * 3].height = 8        # espaco entre linhas de cards

    linha_apos_cards = n_linhas_cards * 3 + 1  # primeira linha livre apos os cards

    def escrever_tabela(linha_inicio, titulo_col_a, titulo_col_b, itens, formatos=None, col_inicio=1):
        ws.cell(row=linha_inicio, column=col_inicio, value=titulo_col_a)
        ws.cell(row=linha_inicio, column=col_inicio + 1, value=titulo_col_b)
        for col in (col_inicio, col_inicio + 1):
            c = ws.cell(row=linha_inicio, column=col)
            c.font = fonte_header
            c.fill = fill_header
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda

        row = linha_inicio + 1
        for i, item in enumerate(itens):
            rotulo, valor = item[0], item[1]
            fmt = item[2] if len(item) > 2 else (formatos[i] if formatos else None)
            c_rot = ws.cell(row=row, column=col_inicio, value=rotulo)
            c_val = ws.cell(row=row, column=col_inicio + 1, value=valor)
            if fmt:
                c_val.number_format = fmt
            c_rot.border = borda
            c_val.border = borda
            c_rot.font = Font(color=BRANCO)
            c_val.font = Font(color=BRANCO)
            c_rot.fill = PatternFill(start_color=BG_CARD, end_color=BG_CARD, fill_type="solid")
            c_val.fill = PatternFill(start_color=BG_CARD, end_color=BG_CARD, fill_type="solid")
            c_rot.alignment = Alignment(horizontal="left", vertical="center")
            c_val.alignment = Alignment(horizontal="center", vertical="center")
            row += 1
        return row - 1  # ultima linha escrita

    # Tabelas auxiliares (STATUS/PRAZO/CATEGORIA/LIDER) ficam a direita dos
    # graficos, colunas N/O (14/15), empilhadas verticalmente
    COL_TABELAS = 14  # N
    linha_tabelas = 1

    # --- Tabela 1: Status (para o grafico de pizza) ---
    linha_status_header = linha_tabelas
    ultima_linha_status = escrever_tabela(linha_status_header, "STATUS", "QUANTIDADE",
                                            resumo["contagens"], col_inicio=COL_TABELAS)

    # --- Tabela 2: Prazo (para o donut) ---
    linha_prazo_header = ultima_linha_status + 2
    itens_prazo_fmt = [(r, v, "0.00%") for r, v in resumo["prazo"]]
    ultima_linha_prazo = escrever_tabela(linha_prazo_header, "PRAZO", "PERCENTUAL", itens_prazo_fmt,
                                           col_inicio=COL_TABELAS)

    # --- Tabela 3: Atrasados por categoria (para o grafico de barras) ---
    linha_categoria_header = ultima_linha_prazo + 2
    itens_categoria = resumo["atrasados_categoria"] or [("(nenhum atraso)", 0)]
    ultima_linha_categoria = escrever_tabela(linha_categoria_header, "CATEGORIA", "ATRASADOS", itens_categoria,
                                               col_inicio=COL_TABELAS)

    # --- Tabela 4: Projetos por lider tecnico ---
    linha_lider_header = ultima_linha_categoria + 2
    itens_lider = resumo["projetos_por_lider"] or [("(sem líder)", 0)]
    ultima_linha_lider = escrever_tabela(linha_lider_header, "LÍDER TÉCNICO", "NÚMERO DE PROJETOS", itens_lider,
                                           col_inicio=COL_TABELAS)

    ws.column_dimensions[get_column_letter(COL_TABELAS)].width = 26
    ws.column_dimensions[get_column_letter(COL_TABELAS + 1)].width = 16

    # Paleta dark/teal para os graficos (consistente com os cards)
    PALETA = [TEAL, AZUL, LARANJA, AMARELO, "9B8CF2", "6B7280", TEAL_CLARO]

    def _escurecer_grafico(chart, titulo_cor=BRANCO):
        """Aplica fundo escuro ao chart space e plot area, e texto claro
        ao titulo/eixos/legenda, para combinar com o tema dark da aba."""
        chart.graphical_properties = GraphicalProperties(solidFill=BG_CARD)
        try:
            chart.plot_area.graphicalProperties = GraphicalProperties(solidFill=BG_CARD)
        except Exception:
            pass
        if chart.title is not None:
            try:
                for para in chart.title.tx.rich.p:
                    for run in para.r:
                        run.rPr = CharacterProperties(solidFill=titulo_cor, b=True)
            except Exception:
                pass
        if chart.legend is not None:
            chart.legend.textProperties = RichText(
                p=[Paragraph(pPr=ParagraphProperties(
                    defRPr=CharacterProperties(solidFill=CINZA_TEXTO)))]
            )
        # Eixos (apenas BarChart/LineChart possuem x_axis/y_axis)
        for axis_attr in ("x_axis", "y_axis"):
            axis = getattr(chart, axis_attr, None)
            if axis is not None:
                axis.txPr = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(
                        defRPr=CharacterProperties(solidFill=CINZA_TEXTO)))]
                )

    # --- Grafico 1: Pizza - distribuicao de status (com rotulos de dados) ---
    pie = PieChart()
    pie.title = "Distribuição de Status dos Projetos"
    # Grafico de pizza exclui a linha "TOTAL" (soma, nao e' uma categoria)
    ultima_linha_pizza = ultima_linha_status - 1
    data = Reference(ws, min_col=COL_TABELAS + 1, min_row=linha_status_header, max_row=ultima_linha_pizza)
    labels = Reference(ws, min_col=COL_TABELAS, min_row=linha_status_header + 1, max_row=ultima_linha_pizza)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    for i in range(ultima_linha_pizza - linha_status_header):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = PALETA[i % len(PALETA)]
        pie.series[0].data_points.append(pt)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.numFmt = "0.00%"
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = False
    pie.dataLabels.showSerName = False
    pie.dataLabels.showLegendKey = False
    pie.legend.position = "b"
    pie.legend.overlay = False
    pie.dataLabels.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(solidFill=BRANCO, b=True)))]
    )

    _escurecer_grafico(pie)
    # Layout fixo: graficos lado a lado, abaixo dos cards (linha 8)
    # Pizza (A8), Donut (J8), Barras (S8) — nunca se sobrepoe
    LINHA_CHARTS = linha_apos_cards  # sempre logo apos os cards
    pie.height = 10
    pie.width = 14
    ws.add_chart(pie, f"A{LINHA_CHARTS}")

    # --- Grafico 2: Donut - % efetivados no prazo vs fora do prazo ---
    donut = PieChart()
    donut.title = "Efetivados: No Prazo x Fora do Prazo"
    try:
        donut.type = "doughnut"
    except Exception:
        pass

    data_prazo = Reference(ws, min_col=COL_TABELAS + 1, min_row=linha_prazo_header, max_row=ultima_linha_prazo)
    labels_prazo = Reference(ws, min_col=COL_TABELAS, min_row=linha_prazo_header + 1, max_row=ultima_linha_prazo)
    donut.add_data(data_prazo, titles_from_data=True)
    donut.set_categories(labels_prazo)

    cores_prazo = [TEAL, LARANJA]  # teal (no prazo) / laranja (fora do prazo)
    for i, cor in enumerate(cores_prazo):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = cor
        donut.series[0].data_points.append(pt)

    donut.dataLabels = DataLabelList()
    donut.dataLabels.showPercent = True
    donut.dataLabels.numFmt = "0.00%"
    donut.dataLabels.showVal = False
    donut.dataLabels.showCatName = False
    donut.dataLabels.showSerName = False
    donut.dataLabels.showLegendKey = False
    donut.legend.position = "b"
    donut.legend.overlay = False
    donut.dataLabels.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(solidFill=BRANCO, b=True)))]
    )

    _escurecer_grafico(donut)
    donut.height = 10
    donut.width = 14
    ws.add_chart(donut, f"J{LINHA_CHARTS}")

    # --- Grafico 3: Barras horizontais - atrasados por categoria ---
    bar = BarChart()
    bar.type = "bar"  # horizontal
    bar.title = "Projetos Atrasados por Categoria (Efetivados + Em Aberto)"
    bar.style = 10

    data_cat = Reference(ws, min_col=COL_TABELAS + 1, min_row=linha_categoria_header, max_row=ultima_linha_categoria)
    labels_cat = Reference(ws, min_col=COL_TABELAS, min_row=linha_categoria_header + 1, max_row=ultima_linha_categoria)
    bar.add_data(data_cat, titles_from_data=True)
    bar.set_categories(labels_cat)

    serie = bar.series[0]
    serie.graphicalProperties.solidFill = LARANJA
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    bar.dataLabels.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(solidFill=BRANCO, b=True)))]
    )
    bar.legend = None

    _escurecer_grafico(bar)
    bar.height = 10
    bar.width = 14
    ws.add_chart(bar, f"S{LINHA_CHARTS}")



COL_DIAS = 2  # B


def recalcular_dias(ws, ultima_linha):
    """
    EN: Recomputes the DIAS column in Python instead of trusting the cached
    formula value (which can be stale, e.g. if it depends on TODAY()/HOJE()).

    Logica (equivalente a Q19 = SE(ETEXTO(Q17); SE(Q17="PRELIMINAR"; G17-J20; "n/a"); K20)):
    - Se DEP (col C) e' uma data (projeto efetivado): DIAS = ESTIMATIVA - DEP
      (negativo = atrasado, positivo = adiantado — mesma convencao do caso PRELIMINAR)
    - Se DEP == "PRELIMINAR" (ainda nao efetivado): DIAS = ESTIMATIVA - HOJE()
      (dias restantes; negativo = estimativa ja vencida / atrasado)
    - Outros status textuais (CONCEITO, CANCELADO, etc.): mantem 'n/a'
    """
    hoje = datetime.date.today()
    for row in range(LINHA_PRIMEIRA_DADO, ultima_linha + 1):
        dep_val = ws.cell(row=row, column=COL_DEP).value
        estimativa = ws.cell(row=row, column=COL_ESTIMATIVA).value
        est_date = estimativa.date() if isinstance(estimativa, datetime.datetime) else estimativa

        if isinstance(dep_val, (datetime.date, datetime.datetime)):
            dep_date = dep_val.date() if isinstance(dep_val, datetime.datetime) else dep_val
            if isinstance(est_date, datetime.date):
                novo_dias = (est_date - dep_date).days
            else:
                novo_dias = None
        elif isinstance(dep_val, str) and dep_val.strip().upper() == "PRELIMINAR":
            if isinstance(est_date, datetime.date):
                novo_dias = (est_date - hoje).days
            else:
                novo_dias = None
        else:
            novo_dias = "n/a"

        ws.cell(row=row, column=COL_DIAS, value=novo_dias)


def _formatar_aba_projetos(ws_dest, colunas_relevantes, linhas_dados, col_assunto_local=None,
                             col_dias_local=None, colunas_data_local=None):
    """
    EN: Shared formatter for "Projetos Efetivados" and "Projetos Preliminares".
    Applies Metalfrio header (dark blue + emerald green), zebra striping,
    borders, ASSUNTO left+middle alignment, auto-filter and freeze panes.
    """
    n_cols = len(colunas_relevantes)
    ultima_linha = 1 + len(linhas_dados)

    fonte_header = Font(bold=True, color=VERDE_METALFRIO)
    fill_header = PatternFill(start_color=AZUL_METALFRIO, end_color=AZUL_METALFRIO, fill_type="solid")
    fill_zebra = PatternFill(start_color="EEF3FA", end_color="EEF3FA", fill_type="solid")
    fina = Side(style="thin", color="BFBFBF")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)

    for j, (_, titulo) in enumerate(colunas_relevantes, start=1):
        cell = ws_dest.cell(row=1, column=j, value=titulo)
        cell.font = fonte_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda

    for i, linha in enumerate(linhas_dados, start=2):
        zebra = (i % 2 == 0)
        for j, valor in enumerate(linha, start=1):
            cell = ws_dest.cell(row=i, column=j, value=valor)
            cell.border = borda
            if col_assunto_local and j == col_assunto_local:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if colunas_data_local and j in colunas_data_local:
                cell.number_format = "dd/mm/yy"
            if zebra:
                cell.fill = fill_zebra

    if ultima_linha >= 2:
        ws_dest.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{ultima_linha}"
        ws_dest.freeze_panes = "A2"

        if col_dias_local:
            from openpyxl.formatting.rule import CellIsRule
            rng = f"{get_column_letter(col_dias_local)}2:{get_column_letter(col_dias_local)}{ultima_linha}"
            ws_dest.conditional_formatting.add(
                rng,
                CellIsRule(operator="lessThan", formula=["0"],
                           fill=PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
                           font=Font(color="782818"))
            )
            ws_dest.conditional_formatting.add(
                rng,
                CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                           fill=PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
                           font=Font(color="286428"))
            )

    for col, (_, titulo) in enumerate(colunas_relevantes, start=1):
        ws_dest.column_dimensions[get_column_letter(col)].width = max(12, len(titulo) + 2)


def criar_aba_efetivados(wb_out, ws_origem, ultima_linha):
    """EN: Creates "Projetos Efetivados" sheet — only rows where DEP is a date."""
    colunas = [
        (1,  "PROJETO"),
        (3,  "DATA EFETIVAÇÃO"),
        (15, "SOLICITAÇÃO"),
        (16, "ESTIMATIVA"),
        (2,  "DIAS (ATRASO)"),
        (13, "ASSUNTO"),
        (8,  "LÍDER TÉCNICO"),
        (7,  "CLASSIFICAÇÃO"),
        (9,  "CATEGORIA"),
        (5,  "LOGIN"),
    ]
    ws_ef = wb_out.create_sheet("Projetos Efetivados")
    linhas = []
    for row in range(LINHA_PRIMEIRA_DADO, ultima_linha + 1):
        dep_val = ws_origem.cell(row=row, column=COL_DEP).value
        if not isinstance(dep_val, (datetime.date, datetime.datetime)):
            continue
        linhas.append([ws_origem.cell(row=row, column=col).value for col, _ in colunas])
    _formatar_aba_projetos(ws_ef, colunas, linhas,
                            col_assunto_local=6, col_dias_local=5,
                            colunas_data_local={2, 3, 4})
    print(f"Aba 'Projetos Efetivados': {len(linhas)} projeto(s)")


def criar_aba_preliminares(wb_out, ws_origem, ultima_linha):
    """EN: Creates "Projetos Preliminares" sheet — all projects."""
    colunas = [
        (1,  "PROJETO"),
        (3,  "DEP"),
        (2,  "DIAS"),
        (16, "ESTIMATIVA"),
        (15, "SOLICITAÇÃO"),
        (13, "ASSUNTO"),
        (8,  "LÍDER TÉCNICO"),
        (7,  "CLASSIFICAÇÃO"),
        (9,  "CATEGORIA"),
        (10, "PROGRAMA"),
        (5,  "LOGIN"),
    ]
    ws_pre = wb_out.create_sheet("Projetos Preliminares")
    linhas = []
    for row in range(LINHA_PRIMEIRA_DADO, ultima_linha + 1):
        linhas.append([ws_origem.cell(row=row, column=col).value for col, _ in colunas])
    _formatar_aba_projetos(ws_pre, colunas, linhas,
                            col_assunto_local=6, col_dias_local=3,
                            colunas_data_local={4, 5})
    print(f"Aba 'Projetos Preliminares': {len(linhas)} projeto(s)")


def aplicar_formatacao(ws, ultima_linha, cabecalhos, total_colunas, colunas_data):
    # Cabecalho: azul Metalfrio + texto verde esmeralda
    fonte_header = Font(bold=True, color=VERDE_METALFRIO)
    fill_header = PatternFill(start_color=AZUL_METALFRIO, end_color=AZUL_METALFRIO, fill_type="solid")
    for col in range(1, total_colunas + 1):
        cell = ws.cell(row=LINHA_HEADER, column=col)
        cell.font = fonte_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if ultima_linha < LINHA_PRIMEIRA_DADO:
        ws.auto_filter.ref = f"A{LINHA_HEADER}:{get_column_letter(total_colunas)}{LINHA_HEADER}"
        ajustar_largura_colunas(ws, cabecalhos)
        return

    # Zebra striping: linhas alternadas com fundo levemente azulado
    fill_zebra = PatternFill(start_color="EEF3FA", end_color="EEF3FA", fill_type="solid")

    for row in range(LINHA_PRIMEIRA_DADO, ultima_linha + 1):
        ws.cell(row=row, column=COL_STATUS_PCT).number_format = "0.00%"
        ws.cell(row=row, column=COL_DEP).number_format = "dd/mm/yy"
        for col in colunas_data:
            if col <= total_colunas:
                ws.cell(row=row, column=col).number_format = "dd/mm/yy"

        zebra = (row % 2 == 0)
        for col in range(1, total_colunas + 1):
            cell = ws.cell(row=row, column=col)
            # ASSUNTO: alinhar à esquerda e ao meio
            if col == COL_ASSUNTO:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if zebra:
                cell.fill = fill_zebra

    fina = Side(style="thin", color="BFBFBF")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)
    for row in range(LINHA_HEADER, ultima_linha + 1):
        for col in range(1, total_colunas + 1):
            ws.cell(row=row, column=col).border = borda

    ws.auto_filter.ref = (
        f"A{LINHA_HEADER}:{get_column_letter(total_colunas)}{ultima_linha}"
    )

    ws.freeze_panes = ws.cell(row=LINHA_PRIMEIRA_DADO, column=1)

    ajustar_largura_colunas(ws, cabecalhos)


def ajustar_largura_colunas(ws, cabecalhos):
    for col, titulo in enumerate(cabecalhos, start=1):
        largura = max(10, len(titulo) + 2)
        ws.column_dimensions[get_column_letter(col)].width = largura




# =====================================================================
# EXTRACAO DE ITENS DI + CRUZAMENTO PPAP - logica adaptada de extrair_itens_di.py
# =====================================================================

DI_PASTA_ORIGEM = r"C:\Projetos"
DI_PASTA_DESTINO = str(Path.home() / "Downloads")
DI_EXTENSOES_VALIDAS = {".xls", ".xlsx", ".xlsm"}

TEXTO_TABELA_COMPONENTES = "RELACAO DE COMPONENTES"
DI_TEXTO_TABELA_PRODUTOS = "CODIGO PRODUTO ACABADO"  # usada so para nao confundir titulos

# Celula onde fica o numero/identificador do projeto (igual ao script de
# Projetos Preliminares)
CELULA_PROJETO = ("Controle de Projeto", "Q15")

# Cabecalhos da tabela "RELACAO DE COMPONENTES" (conforme planilha real)
# nome normalizado -> nome de exibicao na saida
COLUNAS_COMPONENTES = {
    "ITEM": "ITEM",
    "CODIGO": "CODIGO",
    "DESENHO": "DESENHO",
    "LOGIN": "LOGIN",
    "ANT": "ANT",
    "REV": "REV",
    "DENOMINACAO": "DENOMINACAO",
    "TIP": "TIPO",
    "QTDE": "QTDE",
    "UNID": "UNID",
    "ACAO": "ACAO",
    "SIT": "SIT",
    "RIA": "RIA",
    "PLC/RFQ": "PLC/RFQ",
    "APLICACAO": "APLICACAO",
}

# Colunas mantidas na saida (alem de PROJETO), na ordem desejada
COLUNAS_SAIDA = [
    "CODIGO", "DESENHO", "ANT", "REV", "DENOMINACAO",
    "TIPO", "QTDE", "ACAO", "RIA", "PLC/RFQ", "APLICACAO",
]




def di_listar_arquivos(pasta):
    arquivos = []
    for entry in os.scandir(pasta):
        if entry.is_file():
            ext = Path(entry.name).suffix.lower()
            if ext in DI_EXTENSOES_VALIDAS and not entry.name.startswith("~$"):
                arquivos.append(entry.path)
    return arquivos


def extrair_itens_di(ws):
    """
    EN: Locates the "RELAÇÃO DE COMPONENTES" table inside the given
    "Controle de Projeto" worksheet (already open) and returns
    (itens, diag), where itens is a list of dicts {coluna_saida: valor}
    for rows where SIT == "DI", and diag is a short diagnostic string
    ("ok", "titulo nao encontrado", etc.).
    """
    itens = []

    TEXTO_NORM = _norm(TEXTO_TABELA_COMPONENTES)
    TEXTO_PRODUTOS_NORM = _norm(DI_TEXTO_TABELA_PRODUTOS)

    # Localiza o titulo "RELACAO DE COMPONENTES" (ignora "CODIGO
    # PRODUTO ACABADO", que e' uma tabela diferente na mesma aba).
    # Titulos ficam sempre na COLUNA B, normalmente nas primeiras 100
    # linhas -> busca restrita a essa unica coluna/janela (muito mais
    # rapido que varrer todas as colunas de 300 linhas).
    linha_titulo = None
    JANELA_BUSCA_TITULO = min(ws.max_row, 150)  # margem de seguranca sobre as ~100 esperadas
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=JANELA_BUSCA_TITULO,
                                                  min_col=2, max_col=2), start=1):
        cell = row[0]
        if cell.value is None:
            continue
        texto_cel = _norm(cell.value)
        if TEXTO_NORM in texto_cel and TEXTO_PRODUTOS_NORM not in texto_cel:
            linha_titulo = row_idx
            break

    if linha_titulo is None:
        # Fallback raro: tabela mais abaixo do que o esperado -> escaneia
        # o resto da coluna B (ainda muito mais rapido que todas as colunas)
        for row_idx, row in enumerate(ws.iter_rows(min_row=JANELA_BUSCA_TITULO + 1,
                                                      max_row=ws.max_row,
                                                      min_col=2, max_col=2),
                                        start=JANELA_BUSCA_TITULO + 1):
            cell = row[0]
            if cell.value is None:
                continue
            texto_cel = _norm(cell.value)
            if TEXTO_NORM in texto_cel and TEXTO_PRODUTOS_NORM not in texto_cel:
                linha_titulo = row_idx
                break

    if linha_titulo is None:
        return itens, "titulo 'RELACAO DE COMPONENTES' nao encontrado"

    # Busca o cabecalho numa janela de algumas linhas abaixo do titulo
    linha_header = None
    mapa_colunas = {}  # nome_normalizado -> col_idx
    for candidata in range(linha_titulo + 1, min(linha_titulo + 6, ws.max_row + 1)):
        mapa_tmp = {}
        for col_idx, cell in enumerate(next(ws.iter_rows(min_row=candidata, max_row=candidata)), start=1):
            texto = _norm(cell.value)
            if texto in COLUNAS_COMPONENTES and texto not in mapa_tmp:
                mapa_tmp[texto] = col_idx
        # Considera valido se achou pelo menos CODIGO e SIT
        if "CODIGO" in mapa_tmp and "SIT" in mapa_tmp:
            linha_header = candidata
            mapa_colunas = mapa_tmp
            break

    if linha_header is None:
        return itens, "cabecalho da tabela (CODIGO/SIT) nao encontrado"

    col_codigo = mapa_colunas.get("CODIGO")
    col_sit = mapa_colunas.get("SIT")
    col_tipo = mapa_colunas.get("TIP")

    # Le todas as linhas de dados ate encontrar CODIGO e ITEM ambos vazios
    # (linha totalmente vazia indica fim da tabela)
    MAX_ITENS_SEGURANCA = 500
    col_item = mapa_colunas.get("ITEM")

    # Contadores de diagnostico (por que linhas foram descartadas)
    n_linhas_dados = 0
    n_sit_di = 0
    n_sit_di_tipo_comp = 0
    n_sit_di_tipo_comp_cod02 = 0

    for i in range(MAX_ITENS_SEGURANCA):
        linha_dado = linha_header + 1 + i
        if linha_dado > ws.max_row:
            break

        valor_codigo = ws.cell(row=linha_dado, column=col_codigo).value if col_codigo else None
        valor_item = ws.cell(row=linha_dado, column=col_item).value if col_item else None

        # Linha vazia (sem ITEM nem CODIGO) -> fim da tabela
        if (valor_codigo is None or str(valor_codigo).strip() == "") and \
           (valor_item is None or str(valor_item).strip() == ""):
            break

        n_linhas_dados += 1

        valor_sit = ws.cell(row=linha_dado, column=col_sit).value if col_sit else None
        if _norm(valor_sit) != "DI":
            continue
        n_sit_di += 1

        valor_tipo = ws.cell(row=linha_dado, column=col_tipo).value if col_tipo else None
        if _norm(valor_tipo) != "COMP":
            continue
        n_sit_di_tipo_comp += 1

        if valor_codigo is None or str(valor_codigo).strip() == "":
            continue  # SIT=DI/TIPO=COMP mas sem codigo -> ignora

        # Filtro adicional: apenas CODIGOs que comecam com "02"
        # (outros prefixos = outras categorias de itens, fora do escopo)
        codigo_str = str(valor_codigo).strip()
        if not codigo_str.startswith("02"):
            continue
        n_sit_di_tipo_comp_cod02 += 1

        item = {}
        for nome_norm, nome_saida in COLUNAS_COMPONENTES.items():
            col_idx = mapa_colunas.get(nome_norm)
            item[nome_saida] = ws.cell(row=linha_dado, column=col_idx).value if col_idx else None
        itens.append(item)

    if itens:
        diag = "ok"
    else:
        diag = (f"tabela encontrada (header linha {linha_header}, {n_linhas_dados} linha(s) de dados) "
                f"- SIT=DI: {n_sit_di}, SIT=DI+TIPO=COMP: {n_sit_di_tipo_comp}, "
                f"+CODIGO inicia '02': {n_sit_di_tipo_comp_cod02}")
    return itens, diag


def di_ler_arquivo(caminho):
    """
    EN: Worker function (must be top-level/picklable for multiprocessing).
    Opens the workbook ONCE (single read_only pass) and extracts both the
    PROJETO identifier and the DI items from "Controle de Projeto".
    Returns (caminho, projeto, itens, diag, erro_ou_None).
    """
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    except Exception as e:
        return (caminho, None, [], None, str(e))

    try:
        if "Controle de Projeto" not in wb.sheetnames:
            return (caminho, None, [], "sem aba 'Controle de Projeto'", None)
        ws = wb["Controle de Projeto"]

        projeto = None
        try:
            _, celula = CELULA_PROJETO
            projeto = ws[celula].value
        except Exception:
            projeto = None

        try:
            itens, diag = extrair_itens_di(ws)
        except Exception as e:
            itens, diag = [], f"erro: {e}"
    finally:
        wb.close()

    return (caminho, projeto, itens, diag, None)


def main_di(pasta_origem=None, pasta_destino=None, log=print, progress=None):
    pasta_origem = pasta_origem or DI_PASTA_ORIGEM
    pasta_destino = pasta_destino or DI_PASTA_DESTINO

    def _progress(current, total, message=""):
        if progress:
            progress(current, total, message)

    log("=" * 70)
    log("Extracao de Itens DI - YYMMDD.xlsx (v1)")
    log("=" * 70)

    if not os.path.isdir(pasta_origem):
        log(f"ERRO: pasta de origem nao encontrada: {pasta_origem}")
        return None

    arquivos = di_listar_arquivos(pasta_origem)
    total = len(arquivos)
    log(f"Encontrados {total} arquivo(s) em {pasta_origem}\n")
    _progress(0, total, "Iniciando...")

    if total == 0:
        log("Nenhum arquivo para processar. Encerrando.")
        return None

    inicio = datetime.datetime.now()

    n_workers = max(1, min(os.cpu_count() or 4, 8))
    log(f"Lendo arquivos em paralelo usando {n_workers} processo(s)...\n")

    # Tabela 1: todos os itens DI, com PROJETO de origem
    todos = []  # list of dicts: {"PROJETO": ..., **item}
    erros = []
    for i, (caminho, projeto, itens, diag, erro) in enumerate(
            _pool_imap(n_workers, di_ler_arquivo, arquivos), start=1):
        nome = os.path.basename(caminho)
        if erro:
            erros.append((nome, erro))
            log(f"[{i}/{total}] ERRO: {nome} -> {erro}")
        else:
            log(f"[{i}/{total}] OK: {nome} -> {len(itens)} item(ns) DI"
                f"{'' if diag == 'ok' or itens else f' ({diag})'}")
            for item in itens:
                linha = {"PROJETO": projeto}
                linha.update(item)
                todos.append(linha)
        _progress(i, total, f"Lendo arquivos... ({i}/{total})")

    tempo_leitura = (datetime.datetime.now() - inicio).total_seconds()
    log(f"\nLeitura concluida em {tempo_leitura:.1f}s "
        f"({total - len(erros)} ok, {len(erros)} com erro)")
    log(f"Total de itens DI encontrados (todos os projetos): {len(todos)}")

    # Tabela 2: deduplicado por CODIGO (1a ocorrencia)
    unicos = []
    vistos = set()
    for linha in todos:
        codigo = linha.get("CODIGO")
        chave = str(codigo).strip() if codigo is not None else None
        if chave is None or chave == "" or chave in vistos:
            continue
        vistos.add(chave)
        unicos.append(linha)

    log(f"Total de itens DI unicos (por CODIGO): {len(unicos)}")

    # --- Monta o workbook de saida ---
    wb_out = openpyxl.Workbook()

    cabecalhos = ["PROJETO"] + COLUNAS_SAIDA

    _escrever_aba_di(wb_out.active, "Itens DI (Todos)", cabecalhos, todos)
    ws2 = wb_out.create_sheet("Itens DI (Unicos)")
    _escrever_aba_di(ws2, None, cabecalhos, unicos, ja_existe=True)

    hoje = datetime.date.today().strftime("%d %m %Y")
    nome_saida = f"Lista de Itens em Desenvolvimento {hoje}.xlsx"
    caminho_saida = os.path.join(pasta_destino, nome_saida)
    wb_out.save(caminho_saida)

    tempo_total = (datetime.datetime.now() - inicio).total_seconds()
    log(f"\nArquivo salvo em: {caminho_saida}")
    log(f"Tempo total: {tempo_total:.1f}s")
    if erros:
        log(f"\n{len(erros)} arquivo(s) com erro (pulados):")
        for nome, msg in erros:
            log(f"  - {nome}: {msg}")
    log("Concluido.")
    _progress(total, total, "Concluido")
    return caminho_saida


def _pool_imap(n_workers, func, arquivos):
    """EN: Thin wrapper so main_di() doesn't need to manage Pool lifecycle
    inline (keeps the structure close to the Projetos Preliminares
    script)."""
    with mp.Pool(processes=n_workers) as pool:
        for resultado in pool.imap_unordered(func, arquivos):
            yield resultado


def _escrever_aba_di(ws, titulo_aba, cabecalhos, linhas, ja_existe=False):
    """EN: Writes header + data rows into ws, applies basic formatting
    (header style, borders, autofilter, frozen header, column widths)."""
    if titulo_aba:
        ws.title = titulo_aba

    fonte_header = Font(bold=True, color="404040")
    fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fina = Side(style="thin", color="BFBFBF")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)

    for j, titulo in enumerate(cabecalhos, start=1):
        c = ws.cell(row=1, column=j, value=titulo)
        c.font = fonte_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda

    for i, linha in enumerate(linhas, start=2):
        for j, cab in enumerate(cabecalhos, start=1):
            c = ws.cell(row=i, column=j, value=linha.get(cab))
            c.border = borda
            c.alignment = Alignment(horizontal="center", vertical="center")

    ultima_linha = len(linhas) + 1
    total_colunas = len(cabecalhos)
    ws.auto_filter.ref = f"A1:{get_column_letter(total_colunas)}{ultima_linha}"
    ws.freeze_panes = "A2"

    for col, titulo in enumerate(cabecalhos, start=1):
        largura = max(10, len(str(titulo)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = largura



# =====================================================================
# CRUZAMENTO ITENS DI x PPAP (MONDAY) - logica adaptada de cruzar_ppap.py
# =====================================================================

DI_PASTA_DESTINO_PADRAO = str(Path.home() / "Downloads")

# Colunas do Monday (aba "qualidade", header dinamico - procurado por
# texto exato em qualquer linha das primeiras ~10)
COL_MONDAY_CODIGO = "Name"
COL_MONDAY_FORNECEDOR = "Fornecedor"
COL_MONDAY_DEP = "DEP"
COL_MONDAY_EMBALAGEM = "Embalagem"
COL_MONDAY_STATUS = "Status"

ABA_MONDAY_PREFERIDA = "qualidade"


def _ler_itens_di(caminho_itens_di, log):
    """EN: Reads the "Itens DI (Unicos)" sheet, returns (cabecalhos, linhas)
    where linhas is a list of dicts {cabecalho: valor}."""
    wb = openpyxl.load_workbook(caminho_itens_di, data_only=True, read_only=True)
    try:
        nome_aba = None
        for s in wb.sheetnames:
            if "UNIC" in _norm(s):
                nome_aba = s
                break
        if nome_aba is None:
            nome_aba = wb.sheetnames[0]
            log(f"AVISO: aba 'Itens DI (Unicos)' nao encontrada, usando '{nome_aba}'")

        ws = wb[nome_aba]
        cabecalhos = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        linhas = []
        for row in ws.iter_rows(min_row=2):
            valores = [c.value for c in row]
            if all(v is None for v in valores):
                continue
            linhas.append(dict(zip(cabecalhos, valores)))
        return cabecalhos, linhas
    finally:
        wb.close()


def _ler_monday(caminho_monday, log):
    """EN: Reads the Monday "qualidade" export, locates the header row
    dynamically (looks for COL_MONDAY_CODIGO in column A), returns a dict
    keyed by normalized CODIGO -> {Fornecedor, DEP, Embalagem, Status}."""
    wb = openpyxl.load_workbook(caminho_monday, data_only=True, read_only=False)
    try:
        nome_aba = ABA_MONDAY_PREFERIDA if ABA_MONDAY_PREFERIDA in wb.sheetnames else wb.sheetnames[0]
        ws = wb[nome_aba]

        max_row = ws.max_row or 10
        linha_header = None
        mapa_colunas = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_row, 10)), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if _norm(cell.value) == _norm(COL_MONDAY_CODIGO):
                    linha_header = row_idx
                    break
            if linha_header:
                break

        if linha_header is None:
            log("ERRO: cabecalho 'Name' nao encontrado na planilha Monday.")
            return {}

        for col_idx, cell in enumerate(next(ws.iter_rows(min_row=linha_header, max_row=linha_header)), start=1):
            texto = _norm(cell.value)
            for alvo in (COL_MONDAY_CODIGO, COL_MONDAY_FORNECEDOR, COL_MONDAY_DEP,
                         COL_MONDAY_EMBALAGEM, COL_MONDAY_STATUS):
                if texto == _norm(alvo):
                    mapa_colunas[alvo] = col_idx

        faltando = [c for c in (COL_MONDAY_CODIGO, COL_MONDAY_FORNECEDOR, COL_MONDAY_DEP,
                                 COL_MONDAY_EMBALAGEM, COL_MONDAY_STATUS) if c not in mapa_colunas]
        if faltando:
            log(f"AVISO: colunas nao encontradas no Monday: {', '.join(faltando)}")

        registros = {}
        col_codigo = mapa_colunas.get(COL_MONDAY_CODIGO)
        for linha_dado in range(linha_header + 1, max_row + 1):
            valor_codigo = ws.cell(row=linha_dado, column=col_codigo).value if col_codigo else None
            if valor_codigo is None or str(valor_codigo).strip() == "":
                continue
            chave = str(valor_codigo).strip()
            registros[chave] = {
                "Fornecedor": ws.cell(row=linha_dado, column=mapa_colunas[COL_MONDAY_FORNECEDOR]).value if COL_MONDAY_FORNECEDOR in mapa_colunas else None,
                "DEP": ws.cell(row=linha_dado, column=mapa_colunas[COL_MONDAY_DEP]).value if COL_MONDAY_DEP in mapa_colunas else None,
                "Embalagem": ws.cell(row=linha_dado, column=mapa_colunas[COL_MONDAY_EMBALAGEM]).value if COL_MONDAY_EMBALAGEM in mapa_colunas else None,
                "Status": ws.cell(row=linha_dado, column=mapa_colunas[COL_MONDAY_STATUS]).value if COL_MONDAY_STATUS in mapa_colunas else None,
            }
        return registros
    finally:
        wb.close()


def main_cruzar(caminho_itens_di, caminho_monday, pasta_destino=None, log=print, progress=None):
    pasta_destino = pasta_destino or DI_PASTA_DESTINO_PADRAO

    def _progress(current, total, message=""):
        if progress:
            progress(current, total, message)

    log("=" * 70)
    log("Cruzamento Itens DI x PPAP (Monday) - YYMMDD.xlsx (v1)")
    log("=" * 70)

    if not os.path.isfile(caminho_itens_di):
        log(f"ERRO: arquivo 'Itens DI' nao encontrado: {caminho_itens_di}")
        return None
    if not os.path.isfile(caminho_monday):
        log(f"ERRO: arquivo Monday nao encontrado: {caminho_monday}")
        return None

    inicio = datetime.datetime.now()
    _progress(0, 3, "Lendo Itens DI...")

    cabecalhos_di, itens_di = _ler_itens_di(caminho_itens_di, log)
    log(f"Itens DI (unicos) lidos: {len(itens_di)}")

    _progress(1, 3, "Lendo planilha Monday (Qualidade)...")
    monday = _ler_monday(caminho_monday, log)
    log(f"Registros lidos do Monday: {len(monday)}")

    _progress(2, 3, "Cruzando dados...")

    # Determina a coluna CODIGO no Itens DI (geralmente 2a coluna, apos PROJETO)
    if "CODIGO" in cabecalhos_di:
        col_codigo_nome = "CODIGO"
    else:
        # fallback: procura qualquer cabecalho que normalize para CODIGO
        col_codigo_nome = next((c for c in cabecalhos_di if _norm(c) == "CODIGO"), None)

    if col_codigo_nome is None:
        log("ERRO: coluna 'CODIGO' nao encontrada na planilha de Itens DI.")
        return None

    colunas_monday_saida = ["Fornecedor", "DEP", "Embalagem", "Status"]
    cabecalhos_saida = list(cabecalhos_di) + colunas_monday_saida + ["PPAP_DISPONIVEL"]

    cruzamento = []
    sem_ppap = []
    nao_encontrados = []

    for item in itens_di:
        codigo = item.get(col_codigo_nome)
        chave = str(codigo).strip() if codigo is not None else ""

        linha = dict(item)
        registro_monday = monday.get(chave)

        if registro_monday is None:
            for c in colunas_monday_saida:
                linha[c] = None
            linha["PPAP_DISPONIVEL"] = "NAO"
            cruzamento.append(linha)
            nao_encontrados.append(linha)
            continue

        for c in colunas_monday_saida:
            linha[c] = registro_monday.get(c)

        fornecedor = registro_monday.get("Fornecedor")
        ppap_disponivel = "SIM" if fornecedor and str(fornecedor).strip() else "NAO"
        linha["PPAP_DISPONIVEL"] = ppap_disponivel
        cruzamento.append(linha)

        if ppap_disponivel == "NAO":
            sem_ppap.append(linha)

    com_ppap = sum(1 for l in cruzamento if l["PPAP_DISPONIVEL"] == "SIM")
    log(f"\nCruzamento: {len(cruzamento)} itens")
    log(f"  - Com PPAP disponivel: {com_ppap}")
    log(f"  - Sem PPAP (encontrado no Monday, mas fornecedor vazio): {len(sem_ppap)}")
    log(f"  - Nao encontrados no Monday: {len(nao_encontrados)}")

    # --- Monta workbook de saida ---
    wb_out = openpyxl.Workbook()
    _escrever_aba_ppap(wb_out.active, "Cruzamento", cabecalhos_saida, cruzamento)
    _escrever_aba_ppap(wb_out.create_sheet("Sem PPAP"), None, cabecalhos_saida, sem_ppap)
    _escrever_aba_ppap(wb_out.create_sheet("Nao Encontrados"), None, cabecalhos_saida, nao_encontrados)

    hoje = datetime.date.today().strftime("%y%m%d")
    nome_saida = f"Cruzamento PPAP - {hoje}.xlsx"
    caminho_saida = os.path.join(pasta_destino, nome_saida)
    wb_out.save(caminho_saida)

    tempo_total = (datetime.datetime.now() - inicio).total_seconds()
    log(f"\nArquivo salvo em: {caminho_saida}")
    log(f"Tempo total: {tempo_total:.1f}s")
    log("Concluido.")
    _progress(3, 3, "Concluido")
    return caminho_saida


def _escrever_aba_ppap(ws, titulo_aba, cabecalhos, linhas):
    if titulo_aba:
        ws.title = titulo_aba

    fonte_header = Font(bold=True, color="404040")
    fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fina = Side(style="thin", color="BFBFBF")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)

    fill_sim = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fill_nao = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    col_ppap = None
    for j, titulo in enumerate(cabecalhos, start=1):
        c = ws.cell(row=1, column=j, value=titulo)
        c.font = fonte_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda
        if titulo == "PPAP_DISPONIVEL":
            col_ppap = j

    for i, linha in enumerate(linhas, start=2):
        for j, cab in enumerate(cabecalhos, start=1):
            c = ws.cell(row=i, column=j, value=linha.get(cab))
            c.border = borda
            c.alignment = Alignment(horizontal="center", vertical="center")
            if col_ppap and j == col_ppap:
                c.fill = fill_sim if c.value == "SIM" else fill_nao

    ultima_linha = len(linhas) + 1
    total_colunas = len(cabecalhos)
    ws.auto_filter.ref = f"A1:{get_column_letter(total_colunas)}{ultima_linha}"
    ws.freeze_panes = "A2"

    for col, titulo in enumerate(cabecalhos, start=1):
        largura = max(10, len(str(titulo)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = largura


def executar_gui():
    """
    EN: Unified tkinter GUI (dark theme, purple/blue accent) with 3 tabs:
      - "Projetos Preliminares": scans C:\\Projetos, generates the KPI
        dashboard workbook
      - "Extrair Itens DI": scans C:\\Projetos, extracts DI/COMP items
      - "Cruzar PPAP": cross-references Itens DI (Unicos) with Monday
    Each tab has its own pickers, progress bar, status line and
    collapsible log. Processing runs in background threads so the window
    stays responsive.
    """
    import ttkbootstrap as ttk
    from tkinter import filedialog, messagebox, scrolledtext
    import threading

    # ================= JANELA PRINCIPAL (TEMA ESCURO) =================
    root = ttk.Window(themename="cyborg")
    root.title("DNI - Ferramentas | Metalfrio")
    root.geometry("700x380")
    root.resizable(True, True)

    # Acento roxo/azul customizado (estilo da referencia enviada)
    ACCENT = "#6C5CE7"
    ACCENT_HOVER = "#5849c2"
    style = ttk.Style()
    for state_map_name in ("Accent.TButton",):
        style.configure(state_map_name, background=ACCENT, foreground="#FFFFFF",
                        borderwidth=0, focusthickness=0, padding=8)
        style.map(state_map_name,
                  background=[("active", ACCENT_HOVER), ("disabled", "#3a3a4a")],
                  foreground=[("disabled", "#999999")])

    style.configure("Accent.Horizontal.TProgressbar", background=ACCENT,
                    troughcolor="#2b2b3a", borderwidth=0)

    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True, padx=10, pady=(10, 0))

    aba_proj = ttk.Frame(notebook)
    aba_di = ttk.Frame(notebook)
    aba_ppap = ttk.Frame(notebook)
    notebook.add(aba_proj, text="  Projetos Preliminares  ")
    notebook.add(aba_di, text="  Extrair Itens DI  ")
    notebook.add(aba_ppap, text="  Cruzar PPAP  ")

    # ============================================================
    # Helpers compartilhados (log + progress + collapsible details)
    # ============================================================
    def montar_secao_progresso(parent):
        """EN: Builds the shared bottom section (progress bar, status
        label, collapsible log) for a tab. Returns (pbar, lbl_status,
        get_log_widget, log_fn, progress_fn, frame_botao)."""
        frame_botao = ttk.Frame(parent)
        frame_botao.pack(fill="x", padx=12, pady=(15, 5))
        lbl_status = ttk.Label(frame_botao, text="Pronto.", anchor="w")

        frame_progress = ttk.Frame(parent)
        frame_progress.pack(fill="x", padx=12, pady=(0, 8))
        pbar = ttk.Progressbar(frame_progress, orient="horizontal", mode="determinate",
                                style="Accent.Horizontal.TProgressbar")
        pbar.pack(fill="x")

        frame_detalhes = ttk.Frame(parent)
        frame_detalhes.pack(fill="both", expand=True, padx=12, pady=(0, 5))

        state = {"log_widget": None, "visivel": False}

        def alternar_detalhes():
            if state["visivel"]:
                if state["log_widget"] is None:
                    state["log_widget"] = scrolledtext.ScrolledText(
                        frame_detalhes, state="disabled",
                        font=("Consolas", 8), height=9,
                        background="#1a1a24", foreground="#d0d0d8",
                        insertbackground="#d0d0d8", borderwidth=0
                    )
                state["log_widget"].pack(fill="both", expand=True, pady=(5, 0))
                btn_detalhes.config(text="▲ Ocultar detalhes")
            else:
                if state["log_widget"] is not None:
                    state["log_widget"].pack_forget()
                btn_detalhes.config(text="▼ Mostrar detalhes")

        def toggle_detalhes():
            state["visivel"] = not state["visivel"]
            alternar_detalhes()

        btn_detalhes = ttk.Button(parent, text="▼ Mostrar detalhes",
                                   bootstyle="link", command=toggle_detalhes)
        btn_detalhes.pack(side="bottom", pady=(0, 2))

        def log(msg):
            def _append():
                w = state["log_widget"]
                if w is not None:
                    w.configure(state="normal")
                    w.insert("end", str(msg) + "\n")
                    w.see("end")
                    w.configure(state="disabled")
            root.after(0, _append)

        def progress(current, total, message=""):
            def _update():
                if total > 0:
                    pbar.configure(maximum=total, value=current)
                if message:
                    lbl_status.config(text=message)
            root.after(0, _update)

        def limpar_log():
            w = state["log_widget"]
            if w is not None:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.configure(state="disabled")

        return pbar, lbl_status, limpar_log, log, progress, frame_botao

    def montar_seletor_pasta(parent, rotulo, valor_padrao):
        frame = ttk.Frame(parent)
        frame.pack(fill="x", padx=12, pady=5)
        ttk.Label(frame, text=rotulo, width=26, anchor="w").pack(side="left")
        var = ttk.StringVar(value=valor_padrao)
        ttk.Entry(frame, textvariable=var).pack(side="left", fill="x", expand=True, padx=5)

        def escolher():
            d = filedialog.askdirectory(initialdir=var.get() or "/")
            if d:
                var.set(d)

        ttk.Button(frame, text="...", command=escolher, width=3,
                   bootstyle="secondary-outline").pack(side="left")
        return var

    def montar_seletor_arquivo(parent, rotulo, titulo_dialogo):
        frame = ttk.Frame(parent)
        frame.pack(fill="x", padx=12, pady=5)
        ttk.Label(frame, text=rotulo, width=26, anchor="w").pack(side="left")
        var = ttk.StringVar(value="")
        ttk.Entry(frame, textvariable=var).pack(side="left", fill="x", expand=True, padx=5)

        def escolher():
            f = filedialog.askopenfilename(
                title=titulo_dialogo,
                filetypes=[("Excel", "*.xlsx *.xlsm *.xls")]
            )
            if f:
                var.set(f)

        ttk.Button(frame, text="...", command=escolher, width=3,
                   bootstyle="secondary-outline").pack(side="left")
        return var

    # ================= ABA 1: PROJETOS PRELIMINARES =================
    var_origem1 = montar_seletor_pasta(aba_proj, "Pasta de origem (C:\\Projetos):", PASTA_ORIGEM)
    var_destino1 = montar_seletor_pasta(aba_proj, "Pasta de destino (Downloads):", PASTA_DESTINO)

    pbar1, lbl_status1, limpar_log1, log1, progress1, frame_botao1 = montar_secao_progresso(aba_proj)
    btn_proj = ttk.Button(frame_botao1, text="Gerar", width=18, style="Accent.TButton")
    btn_proj.pack(side="left")
    lbl_status1.pack(side="left", padx=10, fill="x", expand=True)

    def rodar_proj():
        btn_proj.config(state="disabled")
        lbl_status1.config(text="Iniciando...")
        pbar1.configure(value=0)
        limpar_log1()

        def trabalho():
            try:
                caminho_saida = main(
                    pasta_origem=var_origem1.get(),
                    pasta_destino=var_destino1.get(),
                    log=log1,
                    progress=progress1,
                )
            except Exception as e:
                log1(f"\nERRO GERAL: {e}")
                caminho_saida = None

            def finalizar():
                btn_proj.config(state="normal")
                lbl_status1.config(text="Pronto." if caminho_saida else "Falhou.")
                if caminho_saida:
                    messagebox.showinfo("Concluído", f"Arquivo gerado com sucesso:\n\n{caminho_saida}")
                else:
                    messagebox.showwarning("Atenção",
                        "Processamento finalizado sem gerar arquivo.\n"
                        "Clique em 'Mostrar detalhes' e gere novamente para ver o log.")

            root.after(0, finalizar)

        threading.Thread(target=trabalho, daemon=True).start()

    btn_proj.config(command=rodar_proj)

    # ================= ABA 2: EXTRAIR ITENS DI =================
    var_origem2 = montar_seletor_pasta(aba_di, "Pasta de origem (C:\\Projetos):", DI_PASTA_ORIGEM)
    var_destino2 = montar_seletor_pasta(aba_di, "Pasta de destino (Downloads):", DI_PASTA_DESTINO)

    pbar2, lbl_status2, limpar_log2, log2, progress2, frame_botao2 = montar_secao_progresso(aba_di)
    btn_di = ttk.Button(frame_botao2, text="Extrair Itens DI", width=18, style="Accent.TButton")
    btn_di.pack(side="left")
    lbl_status2.pack(side="left", padx=10, fill="x", expand=True)

    def rodar_di():
        btn_di.config(state="disabled")
        lbl_status2.config(text="Iniciando...")
        pbar2.configure(value=0)
        limpar_log2()

        def trabalho():
            try:
                caminho_saida = main_di(
                    pasta_origem=var_origem2.get(),
                    pasta_destino=var_destino2.get(),
                    log=log2,
                    progress=progress2,
                )
            except Exception as e:
                log2(f"\nERRO GERAL: {e}")
                caminho_saida = None

            def finalizar():
                btn_di.config(state="normal")
                lbl_status2.config(text="Pronto." if caminho_saida else "Falhou.")
                if caminho_saida:
                    messagebox.showinfo("Concluído", f"Arquivo gerado com sucesso:\n\n{caminho_saida}")
                else:
                    messagebox.showwarning("Atenção",
                        "Processamento finalizado sem gerar arquivo.\n"
                        "Clique em 'Mostrar detalhes' e gere novamente para ver o log.")

            root.after(0, finalizar)

        threading.Thread(target=trabalho, daemon=True).start()

    btn_di.config(command=rodar_di)

    # ================= ABA 3: CRUZAR PPAP =================
    var_di_file = montar_seletor_arquivo(aba_ppap, "Arquivo 'Itens DI - ...xlsx':",
                                          "Selecione o arquivo 'Itens DI - ...xlsx'")
    var_monday = montar_seletor_arquivo(aba_ppap, "Arquivo Monday (Qualidade):",
                                         "Selecione a planilha exportada do Monday (Qualidade)")
    var_destino3 = montar_seletor_pasta(aba_ppap, "Pasta de destino (Downloads):", DI_PASTA_DESTINO_PADRAO)

    pbar3, lbl_status3, limpar_log3, log3, progress3, frame_botao3 = montar_secao_progresso(aba_ppap)
    btn_ppap = ttk.Button(frame_botao3, text="Cruzar", width=18, style="Accent.TButton")
    btn_ppap.pack(side="left")
    lbl_status3.pack(side="left", padx=10, fill="x", expand=True)

    def rodar_ppap():
        if not var_di_file.get() or not var_monday.get():
            messagebox.showwarning("Atenção", "Selecione os dois arquivos (Itens DI e Monday).")
            return

        btn_ppap.config(state="disabled")
        lbl_status3.config(text="Iniciando...")
        pbar3.configure(value=0)
        limpar_log3()

        def trabalho():
            try:
                caminho_saida = main_cruzar(
                    caminho_itens_di=var_di_file.get(),
                    caminho_monday=var_monday.get(),
                    pasta_destino=var_destino3.get(),
                    log=log3,
                    progress=progress3,
                )
            except Exception as e:
                log3(f"\nERRO GERAL: {e}")
                caminho_saida = None

            def finalizar():
                btn_ppap.config(state="normal")
                lbl_status3.config(text="Pronto." if caminho_saida else "Falhou.")
                if caminho_saida:
                    messagebox.showinfo("Concluído", f"Arquivo gerado com sucesso:\n\n{caminho_saida}")
                else:
                    messagebox.showwarning("Atenção",
                        "Processamento finalizado sem gerar arquivo.\n"
                        "Clique em 'Mostrar detalhes' e gere novamente para ver o log.")

            root.after(0, finalizar)

        threading.Thread(target=trabalho, daemon=True).start()

    btn_ppap.config(command=rodar_ppap)

    # --- Rodape: autor (comum as 3 abas) ---
    ttk.Label(root, text=f"Autor: {__author__} | DNI - Metalfrio Solutions",
             font=("Segoe UI", 8), bootstyle="secondary").pack(side="bottom", pady=(0, 5))

    root.mainloop()


if __name__ == "__main__":
    mp.freeze_support()
    executar_gui()
