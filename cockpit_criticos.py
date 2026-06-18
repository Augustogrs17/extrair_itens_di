"""
Cockpit de Itens Críticos — Metalfrio Solutions
Lê exports do TOTVS (ESPL0604 + ESPL0704) e gera Excel consolidado.
Build: pyinstaller --onefile --windowed --icon icon_dni.ico cockpit_criticos.py
"""

import os, sys, threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, date
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Identidade visual Metalfrio ──────────────────────────────────────────────
AZUL   = "0B1D2F"
VERDE  = "1CBABE"
BRANCO = "FFFFFF"
CINZA  = "F2F4F6"

FAROL_MAP = {
    "Vermelho": ("C0392B", "FFFFFF"),
    "Laranja":  ("E67E22", "FFFFFF"),
    "Amarelo":  ("F1C40F", "1A1A1A"),
    "Verde":    ("27AE60", "FFFFFF"),
    "?":        ("BDC3C7", "1A1A1A"),
}
PRIORIDADE = {"Vermelho": 0, "Laranja": 1, "Amarelo": 2, "Verde": 3, "?": 4}

# ─── Utilitários ──────────────────────────────────────────────────────────────

def serial_para_data(v):
    if v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() in ("", "?", "nan"):
        return None
    if isinstance(v, (datetime, date)):
        return v if isinstance(v, date) else v.date()
    try:
        return (datetime(1899, 12, 30) + pd.Timedelta(days=float(v))).date()
    except Exception:
        pass
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def inferir_farol(row):
    f = str(row.get("Farol", "")).strip()
    if f in FAROL_MAP:
        return f
    try:
        c = float(row.get("Cobertura", ""))
        if c <= 0:  return "Vermelho"
        if c <= 7:  return "Laranja"
        if c <= 15: return "Amarelo"
        return "Verde"
    except Exception:
        pass
    try:
        if float(row.get("Saldo", 1)) <= 0:
            return "Vermelho"
    except Exception:
        pass
    return "?"


def bordas():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ─── Leitura ──────────────────────────────────────────────────────────────────

def ler_espl0604(path):
    xl = pd.ExcelFile(path)
    df = xl.parse(xl.sheet_names[0], skiprows=1, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df["Item"] != "Item"].dropna(subset=["Item"]).reset_index(drop=True)
    df["Item"] = df["Item"].str.strip()
    return df


def ler_espl0704(path):
    xl = pd.ExcelFile(path)
    aba = next((s for s in xl.sheet_names if "Itens" in s), xl.sheet_names[0])
    df = xl.parse(aba, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=["Item"])
    df["Item"] = df["Item"].str.strip()
    return df


# ─── Consolidação ─────────────────────────────────────────────────────────────

MAPA_0604 = {
    "Item": "Item", "Descrição": "Descrição", "Saldo WMS": "Saldo WMS",
    "Data Ruptura": "Data Ruptura", "Planejador": "Planejador",
    "Fornecedores": "Fornecedores",
    "Produtos (Onde-se-Usa 30d)": "Onde-Usa 30d",
    "Linhas (Onde-se-Usa 30d)": "Linhas 30d",
    "Produtos (Onde-se-Usa >30d)": "Onde-Usa >30d",
    "Linhas (Onde-se-Usa >30d)": "Linhas >30d",
}
MAPA_0704 = {
    "Item": "Item", "Farol": "Farol", "Saldo": "Saldo",
    "Sld Processo": "Sld Processo", "Sld Em Receb": "Sld Em Receb",
    "Cobertura": "Cobertura", "Cobertura WMS": "Cobertura WMS",
    "Qtd Ordens Atraso": "Qtd OC Atraso", "Qtd Ordens Firmes": "Qtd OC Firmes",
    "Qtd Ordens Planejadas": "Qtd OC Planej",
    "Reservas Confirmadas": "Reservas Conf",
    "Custo Unitário": "Custo Unit",
    "Quant Segur": "Estq Seg", "Tmp Segur": "Tmp Seg",
}


def consolidar(df0604, df0704):
    left  = df0604.rename(columns={k: v for k, v in MAPA_0604.items() if k in df0604.columns})
    right = df0704.rename(columns={k: v for k, v in MAPA_0704.items() if k in df0704.columns})
    left  = left[[c for c in MAPA_0604.values() if c in left.columns]]
    right = right[[c for c in MAPA_0704.values() if c in right.columns]]

    merged = pd.merge(left, right, on="Item", how="outer", suffixes=("_a", "_b"))

    # Descrição unificada
    for sufixo in ("_a", "_b"):
        col = f"Descrição{sufixo}"
        if col in merged.columns:
            merged["Descrição"] = merged.get("Descrição", pd.Series(dtype=str)).combine_first(merged[col])
            merged.drop(columns=[col], inplace=True, errors="ignore")

    merged["Farol"] = merged.apply(inferir_farol, axis=1)
    merged["Data Ruptura"] = merged["Data Ruptura"].apply(serial_para_data)
    merged["Dias p/ Ruptura"] = merged["Data Ruptura"].apply(
        lambda d: (d - date.today()).days if d else None
    )
    merged["_p"] = merged["Farol"].map(PRIORIDADE).fillna(4)
    merged.sort_values(["_p", "Dias p/ Ruptura"], ascending=[True, True],
                       na_position="last", inplace=True)
    merged.drop(columns=["_p"], inplace=True)
    return merged.reset_index(drop=True)


# ─── Exportação Excel ─────────────────────────────────────────────────────────

COLUNAS_SAIDA = [
    "Item", "Descrição", "Farol", "Data Ruptura", "Dias p/ Ruptura",
    "Saldo WMS", "Saldo", "Sld Processo", "Cobertura", "Cobertura WMS",
    "Qtd OC Atraso", "Qtd OC Firmes", "Qtd OC Planej", "Reservas Conf",
    "Fornecedores", "Onde-Usa 30d", "Linhas 30d", "Planejador",
]
LARGURAS = {
    1: 14, 2: 46, 3: 10, 4: 13, 5: 12, 6: 10, 7: 9, 8: 10,
    9: 10, 10: 12, 11: 12, 12: 12, 13: 12, 14: 12,
    15: 30, 16: 22, 17: 22, 18: 24,
}


def _cabecalho_sheet(ws, titulo, subtitulo, n_cols):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    last = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last}1")
    c1 = ws["A1"]
    c1.value = titulo
    c1.font = Font(name="Segoe UI", bold=True, size=14, color=BRANCO)
    c1.fill = PatternFill("solid", fgColor=AZUL)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last}2")
    c2 = ws["A2"]
    c2.value = subtitulo
    c2.font = Font(name="Segoe UI", size=9, color="8A9BAA")
    c2.fill = PatternFill("solid", fgColor=AZUL)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 15


def _header_cols(ws, row, cols):
    for ci, nome in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=nome)
        c.font = Font(name="Segoe UI", bold=True, size=9, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=VERDE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bordas()
    ws.row_dimensions[row].height = 26


def _escrever_linhas(ws, df, cols, start_row):
    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + i
        farol = row.get("Farol", "Verde")
        bg_f, fg_f = FAROL_MAP.get(farol, ("BDC3C7", "1A1A1A"))
        bg_z = CINZA if i % 2 else BRANCO

        for ci, col in enumerate(cols, 1):
            val = row.get(col, "")
            if col == "Data Ruptura" and isinstance(val, date):
                val = val.strftime("%d/%m/%Y")
            elif col == "Dias p/ Ruptura":
                try:
                    val = int(val) if pd.notna(val) else "—"
                except Exception:
                    val = "—"
            elif col in ("Custo Unit", "Cobertura", "Cobertura WMS"):
                try:
                    val = round(float(val), 2) if pd.notna(val) and str(val) not in ("", "nan") else "—"
                except Exception:
                    val = "—"
            else:
                val = str(val).strip() if pd.notna(val) and str(val) != "nan" else ""

            c = ws.cell(row=r, column=ci, value=val)
            c.border = bordas()
            c.alignment = Alignment(vertical="center")
            if col == "Farol":
                c.fill = PatternFill("solid", fgColor=bg_f)
                c.font = Font(name="Segoe UI", size=9, bold=True, color=fg_f)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = PatternFill("solid", fgColor=bg_z)
                c.font = Font(name="Segoe UI", size=9)
        ws.row_dimensions[r].height = 16


def exportar_excel(df, path_saida):
    wb = openpyxl.Workbook()
    cols = [c for c in COLUNAS_SAIDA if c in df.columns]
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ── ABA PRINCIPAL ─────────────────────────────────────────────────────────
    ws_main = wb.active
    ws_main.title = "Resumo Críticos"
    _cabecalho_sheet(ws_main,
                     "COCKPIT DE ITENS CRÍTICOS — METALFRIO SOLUTIONS",
                     f"Gerado em {ts}  |  {len(df)} itens  |  DNI",
                     len(cols))
    _header_cols(ws_main, 3, cols)
    _escrever_linhas(ws_main, df, cols, 4)
    for c, w in LARGURAS.items():
        if c <= len(cols):
            ws_main.column_dimensions[get_column_letter(c)].width = w

    # ── ABAS POR FAROL ────────────────────────────────────────────────────────
    emojis = {"Vermelho": "🔴", "Laranja": "🟠", "Amarelo": "🟡", "Verde": "🟢"}
    for farol_nome in ["Vermelho", "Laranja", "Amarelo", "Verde"]:
        sub = df[df["Farol"] == farol_nome]
        if sub.empty:
            continue
        emoji = emojis.get(farol_nome, "")
        ws = wb.create_sheet(title=f"{emoji} {farol_nome}")
        _cabecalho_sheet(ws,
                         f"ITENS CRÍTICOS — {farol_nome.upper()}",
                         f"{len(sub)} itens  |  {ts}",
                         len(cols))
        _header_cols(ws, 3, cols)
        _escrever_linhas(ws, sub.reset_index(drop=True), cols, 4)
        for c, w in LARGURAS.items():
            if c <= len(cols):
                ws.column_dimensions[get_column_letter(c)].width = w

    # ── ABA KPIs ──────────────────────────────────────────────────────────────
    wk = wb.create_sheet("📊 KPIs")
    wk.sheet_view.showGridLines = False
    wk.merge_cells("A1:B1")
    wk["A1"].value = "INDICADORES — ITENS CRÍTICOS"
    wk["A1"].font = Font(name="Segoe UI", bold=True, size=12, color=BRANCO)
    wk["A1"].fill = PatternFill("solid", fgColor=AZUL)
    wk["A1"].alignment = Alignment(horizontal="left", indent=1, vertical="center")
    wk.row_dimensions[1].height = 26

    def n_dias(x, lim): 
        try: return isinstance(x, (int, float)) and x <= lim
        except: return False

    kpis = [
        ("Total de Itens Críticos", len(df)),
        ("🔴 Vermelho", len(df[df["Farol"] == "Vermelho"])),
        ("🟠 Laranja",  len(df[df["Farol"] == "Laranja"])),
        ("🟡 Amarelo",  len(df[df["Farol"] == "Amarelo"])),
        ("🟢 Verde",    len(df[df["Farol"] == "Verde"])),
        ("Ruptura já ocorrida (< 0 dias)", len(df[df["Dias p/ Ruptura"].apply(lambda x: n_dias(x, -1))])),
        ("Ruptura em ≤ 7 dias",  len(df[df["Dias p/ Ruptura"].apply(lambda x: n_dias(x, 7))])),
        ("Ruptura em ≤ 30 dias", len(df[df["Dias p/ Ruptura"].apply(lambda x: n_dias(x, 30))])),
        ("Saldo WMS = 0", len(df[df.get("Saldo WMS", pd.Series(dtype=str)).apply(
            lambda x: str(x).strip() in ("0", "0.0", ""))])),
    ]
    for i, (label, val) in enumerate(kpis):
        r = i + 2
        ca = wk.cell(row=r, column=1, value=label)
        cb = wk.cell(row=r, column=2, value=val)
        bg = CINZA if i % 2 else BRANCO
        for c in (ca, cb):
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = bordas()
            c.font = Font(name="Segoe UI", size=10)
        cb.font = Font(name="Segoe UI", bold=True, size=11)
        cb.alignment = Alignment(horizontal="center", vertical="center")
        wk.row_dimensions[r].height = 20
    wk.column_dimensions["A"].width = 38
    wk.column_dimensions["B"].width = 14

    wb.save(path_saida)


# ─── Interface Gráfica ────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Cockpit de Itens Críticos — Metalfrio Solutions")
        self.geometry("740x500")
        self.resizable(False, False)
        self.configure(bg=f"#{AZUL}")
        self._set_icon()

        self.path_0604 = tk.StringVar()
        self.path_0704 = tk.StringVar()
        self.path_saida = tk.StringVar(value=os.path.join(
            os.path.expanduser("~"), "Desktop",
            f"Criticos_{datetime.now().strftime('%d%m%y')}.xlsx"
        ))
        self.status  = tk.StringVar(value="Pronto.")
        self.progresso = tk.DoubleVar(value=0)
        self._build()

    def _set_icon(self):
        try:
            base = getattr(sys, "_MEIPASS", os.path.dirname(__file__))
            ico = os.path.join(base, "icon_dni.ico")
            if os.path.exists(ico):
                self.iconbitmap(ico)
        except Exception:
            pass

    def _build(self):
        # Cabeçalho
        tk.Frame(self, bg=f"#{AZUL}", height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=f"#{AZUL}", pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="COCKPIT DE ITENS CRÍTICOS",
                 font=("Segoe UI", 17, "bold"),
                 bg=f"#{AZUL}", fg=f"#{VERDE}").pack()
        tk.Label(hdr, text="Metalfrio Solutions  •  DNI — Desenvolvimento de Novos Itens",
                 font=("Segoe UI", 9), bg=f"#{AZUL}", fg="#5A8FAA").pack(pady=(2, 0))

        # Separador teal
        tk.Frame(self, bg=f"#{VERDE}", height=3).pack(fill="x")

        # Corpo
        body = tk.Frame(self, bg="#EFF2F5", padx=32, pady=22)
        body.pack(fill="both", expand=True)

        self._campo(body, "1.  Arquivo ESPL0604   —   Listagem Itens Críticos",
                    self.path_0604, 0)
        self._campo(body, "2.  Arquivo ESPL0704   —   Demonstrativo Cálculo MRP",
                    self.path_0704, 2)
        self._campo_saida(body, 4)

        # Botão
        btn_frame = tk.Frame(body, bg="#EFF2F5")
        btn_frame.grid(row=6, column=0, columnspan=3, pady=18)
        self.btn = tk.Button(
            btn_frame, text="▶   PROCESSAR E EXPORTAR",
            font=("Segoe UI", 11, "bold"),
            bg=f"#{VERDE}", fg="white",
            activebackground="#139A9D", activeforeground="white",
            relief="flat", padx=28, pady=10, cursor="hand2",
            command=self._iniciar
        )
        self.btn.pack()

        # Progress + status
        ttk.Progressbar(body, variable=self.progresso, maximum=100,
                        mode="determinate", length=660).grid(
            row=7, column=0, columnspan=3, pady=(0, 6))
        tk.Label(body, textvariable=self.status,
                 font=("Segoe UI", 9), bg="#EFF2F5", fg="#555").grid(
            row=8, column=0, columnspan=3)

        # Rodapé
        tk.Frame(self, bg=f"#{AZUL}", height=3).pack(fill="x", side="bottom")
        rod = tk.Frame(self, bg=f"#{AZUL}", pady=5)
        rod.pack(fill="x", side="bottom")
        tk.Label(rod, text="DNI  •  Metalfrio Solutions",
                 font=("Segoe UI", 8), bg=f"#{AZUL}", fg="#3D6880").pack()

    def _campo(self, parent, label, var, row):
        tk.Label(parent, text=label, font=("Segoe UI", 9, "bold"),
                 bg="#EFF2F5", fg=f"#{AZUL}").grid(
            row=row, column=0, columnspan=3, sticky="w", pady=(8, 2))
        tk.Entry(parent, textvariable=var, width=74,
                 font=("Segoe UI", 9), relief="solid", bd=1).grid(
            row=row+1, column=0, columnspan=2, sticky="ew")
        tk.Button(parent, text="📂 Selecionar",
                  font=("Segoe UI", 9), bg=f"#{AZUL}", fg="white",
                  relief="flat", padx=10, cursor="hand2",
                  command=lambda v=var: self._abrir(v)).grid(
            row=row+1, column=2, padx=(6, 0))

    def _campo_saida(self, parent, row):
        tk.Label(parent, text="3.  Arquivo de saída   (.xlsx)",
                 font=("Segoe UI", 9, "bold"),
                 bg="#EFF2F5", fg=f"#{AZUL}").grid(
            row=row, column=0, columnspan=3, sticky="w", pady=(8, 2))
        tk.Entry(parent, textvariable=self.path_saida, width=74,
                 font=("Segoe UI", 9), relief="solid", bd=1).grid(
            row=row+1, column=0, columnspan=2, sticky="ew")
        tk.Button(parent, text="💾 Salvar como",
                  font=("Segoe UI", 9), bg=f"#{AZUL}", fg="white",
                  relief="flat", padx=10, cursor="hand2",
                  command=self._salvar_como).grid(row=row+1, column=2, padx=(6, 0))

    def _abrir(self, var):
        f = filedialog.askopenfilename(
            title="Selecionar arquivo TOTVS",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if f: var.set(f)

    def _salvar_como(self):
        f = filedialog.asksaveasfilename(
            title="Salvar relatório", defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if f: self.path_saida.set(f)

    def _iniciar(self):
        erros = []
        if not self.path_0604.get() or not os.path.exists(self.path_0604.get()):
            erros.append("• Arquivo ESPL0604 não encontrado.")
        if not self.path_0704.get() or not os.path.exists(self.path_0704.get()):
            erros.append("• Arquivo ESPL0704 não encontrado.")
        if not self.path_saida.get():
            erros.append("• Defina o arquivo de saída.")
        if erros:
            messagebox.showerror("Campos obrigatórios", "\n".join(erros))
            return
        self.btn.config(state="disabled")
        threading.Thread(target=self._processar, daemon=True).start()

    def _up(self, pct, msg):
        self.progresso.set(pct)
        self.status.set(msg)
        self.update_idletasks()

    def _processar(self):
        try:
            self._up(10, "Lendo ESPL0604…")
            df0604 = ler_espl0604(self.path_0604.get())

            self._up(35, "Lendo ESPL0704…")
            df0704 = ler_espl0704(self.path_0704.get())

            self._up(55, "Consolidando e calculando farois…")
            df = consolidar(df0604, df0704)

            self._up(78, "Exportando Excel formatado…")
            exportar_excel(df, self.path_saida.get())

            self._up(100, f"✅  Concluído! {len(df)} itens → {os.path.basename(self.path_saida.get())}")
            messagebox.showinfo("Sucesso",
                f"Relatório gerado!\n\n"
                f"Itens consolidados: {len(df)}\n"
                f"🔴 Vermelho: {len(df[df['Farol']=='Vermelho'])}\n"
                f"🟠 Laranja:  {len(df[df['Farol']=='Laranja'])}\n"
                f"🟡 Amarelo:  {len(df[df['Farol']=='Amarelo'])}\n"
                f"🟢 Verde:    {len(df[df['Farol']=='Verde'])}\n\n"
                f"Arquivo: {self.path_saida.get()}")
            try:
                os.startfile(os.path.dirname(self.path_saida.get()))
            except Exception:
                pass
        except Exception as e:
            self._up(0, f"❌  Erro: {e}")
            messagebox.showerror("Erro no processamento", str(e))
        finally:
            self.btn.config(state="normal")


if __name__ == "__main__":
    App().mainloop()
