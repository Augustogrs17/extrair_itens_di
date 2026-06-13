"""
Extrai Itens DI - YYMMDD.xlsx (v1)
==================================
Autor: Augusto G. Silvestrin (silvestrin) - Estagiario Eng. Produto, DNI - Metalfrio Solutions

EN: Scans all project files in a folder, locates the "RELAÇÃO DE
COMPONENTES" table inside the "Controle de Projeto" sheet of each file,
filters rows where SIT == "DI" (Desenvolvimento de Item), and produces a
consolidated spreadsheet with two sheets:
  - "Itens DI (Todos)": every DI row found, tagged with its source PROJETO
  - "Itens DI (Unicos)": deduplicated by CODIGO (first occurrence wins),
    keeping the source PROJETO as reference, for later cross-checking
    against the Monday "Conteudo - Qualidade" PPAP data.

Requisitos: pip install openpyxl
Build: pyinstaller --onefile --windowed --name ExtrairItensDI --version-file version_info.txt extrair_itens_di.py
"""

__author__ = "Augusto G. Silvestrin (silvestrin)"

import os
import sys
import datetime
import multiprocessing as mp
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# =====================================================================
# CONFIGURACAO
# =====================================================================
PASTA_ORIGEM = r"C:\Projetos"
PASTA_DESTINO = str(Path.home() / "Downloads")
EXTENSOES_VALIDAS = {".xls", ".xlsx", ".xlsm"}

TEXTO_TABELA_COMPONENTES = "RELACAO DE COMPONENTES"
TEXTO_TABELA_PRODUTOS = "CODIGO PRODUTO ACABADO"  # usada so para nao confundir titulos

# Celula onde fica o numero/identificador do projeto (igual ao script de
# Projetos Preliminares)
CELULA_PROJETO = ("Controle de Projeto", "Q15")

# Cabecalhos da tabela "RELACAO DE COMPONENTES" (conforme planilha real)
# nome normalizado -> nome de exibicao na saida
COLUNAS_COMPONENTES = {
    "ITEM": "ITEM",
    "CODIGO": "CODIGO",
    "DESENHO": "DESENHO",
    "LOGIN": "LOGIN",
    "ANT": "ANT",
    "REV": "REV",
    "DENOMINACAO": "DENOMINACAO",
    "TIP": "TIPO",
    "QTDE": "QTDE",
    "UNID": "UNID",
    "ACAO": "ACAO",
    "SIT": "SIT",
    "RIA": "RIA",
    "PLC/RFQ": "PLC/RFQ",
    "APLICACAO": "APLICACAO",
}

# Colunas mantidas na saida (alem de PROJETO), na ordem desejada
COLUNAS_SAIDA = [
    "CODIGO", "DESENHO", "ANT", "REV", "DENOMINACAO",
    "TIPO", "QTDE", "ACAO", "RIA", "PLC/RFQ", "APLICACAO",
]


def _norm(texto):
    """EN: Normalize text for robust matching: uppercase, strip accents,
    collapse internal whitespace."""
    if not isinstance(texto, str):
        return ""
    s = texto.strip().upper()
    substituicoes = {
        "Á": "A", "À": "A", "Â": "A", "Ã": "A",
        "É": "E", "È": "E", "Ê": "E",
        "Í": "I", "Ì": "I",
        "Ó": "O", "Ò": "O", "Ô": "O", "Õ": "O",
        "Ú": "U", "Ù": "U",
        "Ç": "C",
    }
    for k, v in substituicoes.items():
        s = s.replace(k, v)
    return " ".join(s.split())  # colapsa espacos multiplos


def listar_arquivos(pasta):
    arquivos = []
    for entry in os.scandir(pasta):
        if entry.is_file():
            ext = Path(entry.name).suffix.lower()
            if ext in EXTENSOES_VALIDAS and not entry.name.startswith("~$"):
                arquivos.append(entry.path)
    return arquivos


def extrair_itens_di(ws):
    """
    EN: Locates the "RELAÇÃO DE COMPONENTES" table inside the given
    "Controle de Projeto" worksheet (already open) and returns
    (itens, diag), where itens is a list of dicts {coluna_saida: valor}
    for rows where SIT == "DI", and diag is a short diagnostic string
    ("ok", "titulo nao encontrado", etc.).
    """
    itens = []

    TEXTO_NORM = _norm(TEXTO_TABELA_COMPONENTES)
    TEXTO_PRODUTOS_NORM = _norm(TEXTO_TABELA_PRODUTOS)

    # Localiza o titulo "RELACAO DE COMPONENTES" (ignora "CODIGO
    # PRODUTO ACABADO", que e' uma tabela diferente na mesma aba)
    linha_titulo = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 300)), start=1):
        for cell in row:
            texto_cel = _norm(cell.value)
            if TEXTO_NORM in texto_cel and TEXTO_PRODUTOS_NORM not in texto_cel:
                linha_titulo = row_idx
                break
        if linha_titulo:
            break

    if linha_titulo is None:
        return itens, "titulo 'RELACAO DE COMPONENTES' nao encontrado"

    # Busca o cabecalho numa janela de algumas linhas abaixo do titulo
    linha_header = None
    mapa_colunas = {}  # nome_normalizado -> col_idx
    for candidata in range(linha_titulo + 1, min(linha_titulo + 6, ws.max_row + 1)):
        mapa_tmp = {}
        for col_idx, cell in enumerate(next(ws.iter_rows(min_row=candidata, max_row=candidata)), start=1):
            texto = _norm(cell.value)
            if texto in COLUNAS_COMPONENTES and texto not in mapa_tmp:
                mapa_tmp[texto] = col_idx
        # Considera valido se achou pelo menos CODIGO e SIT
        if "CODIGO" in mapa_tmp and "SIT" in mapa_tmp:
            linha_header = candidata
            mapa_colunas = mapa_tmp
            break

    if linha_header is None:
        return itens, "cabecalho da tabela (CODIGO/SIT) nao encontrado"

    col_codigo = mapa_colunas.get("CODIGO")
    col_sit = mapa_colunas.get("SIT")

    # Le todas as linhas de dados ate encontrar CODIGO e ITEM ambos vazios
    # (linha totalmente vazia indica fim da tabela)
    MAX_ITENS_SEGURANCA = 500
    col_item = mapa_colunas.get("ITEM")
    for i in range(MAX_ITENS_SEGURANCA):
        linha_dado = linha_header + 1 + i
        if linha_dado > ws.max_row:
            break

        valor_codigo = ws.cell(row=linha_dado, column=col_codigo).value if col_codigo else None
        valor_item = ws.cell(row=linha_dado, column=col_item).value if col_item else None

        # Linha vazia (sem ITEM nem CODIGO) -> fim da tabela
        if (valor_codigo is None or str(valor_codigo).strip() == "") and \
           (valor_item is None or str(valor_item).strip() == ""):
            break

        valor_sit = ws.cell(row=linha_dado, column=col_sit).value if col_sit else None
        if _norm(valor_sit) != "DI":
            continue

        if valor_codigo is None or str(valor_codigo).strip() == "":
            continue  # SIT=DI mas sem codigo -> ignora

        item = {}
        for nome_norm, nome_saida in COLUNAS_COMPONENTES.items():
            col_idx = mapa_colunas.get(nome_norm)
            item[nome_saida] = ws.cell(row=linha_dado, column=col_idx).value if col_idx else None
        itens.append(item)

    diag = "ok" if itens else f"tabela encontrada (header linha {linha_header}) mas 0 itens com SIT=DI"
    return itens, diag


def ler_arquivo(caminho):
    """
    EN: Worker function (must be top-level/picklable for multiprocessing).
    Opens the workbook ONCE (single read_only pass) and extracts both the
    PROJETO identifier and the DI items from "Controle de Projeto".
    Returns (caminho, projeto, itens, diag, erro_ou_None).
    """
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    except Exception as e:
        return (caminho, None, [], None, str(e))

    try:
        if "Controle de Projeto" not in wb.sheetnames:
            return (caminho, None, [], "sem aba 'Controle de Projeto'", None)
        ws = wb["Controle de Projeto"]

        projeto = None
        try:
            _, celula = CELULA_PROJETO
            projeto = ws[celula].value
        except Exception:
            projeto = None

        try:
            itens, diag = extrair_itens_di(ws)
        except Exception as e:
            itens, diag = [], f"erro: {e}"
    finally:
        wb.close()

    return (caminho, projeto, itens, diag, None)


def main(pasta_origem=None, pasta_destino=None, log=print, progress=None):
    pasta_origem = pasta_origem or PASTA_ORIGEM
    pasta_destino = pasta_destino or PASTA_DESTINO

    def _progress(current, total, message=""):
        if progress:
            progress(current, total, message)

    log("=" * 70)
    log("Extracao de Itens DI - YYMMDD.xlsx (v1)")
    log("=" * 70)

    if not os.path.isdir(pasta_origem):
        log(f"ERRO: pasta de origem nao encontrada: {pasta_origem}")
        return None

    arquivos = listar_arquivos(pasta_origem)
    total = len(arquivos)
    log(f"Encontrados {total} arquivo(s) em {pasta_origem}\n")
    _progress(0, total, "Iniciando...")

    if total == 0:
        log("Nenhum arquivo para processar. Encerrando.")
        return None

    inicio = datetime.datetime.now()

    n_workers = max(1, min(os.cpu_count() or 4, 8))
    log(f"Lendo arquivos em paralelo usando {n_workers} processo(s)...\n")

    # Tabela 1: todos os itens DI, com PROJETO de origem
    todos = []  # list of dicts: {"PROJETO": ..., **item}
    erros = []
    for i, (caminho, projeto, itens, diag, erro) in enumerate(
            _pool_imap(n_workers, ler_arquivo, arquivos), start=1):
        nome = os.path.basename(caminho)
        if erro:
            erros.append((nome, erro))
            log(f"[{i}/{total}] ERRO: {nome} -> {erro}")
        else:
            log(f"[{i}/{total}] OK: {nome} -> {len(itens)} item(ns) DI"
                f"{'' if diag == 'ok' or itens else f' ({diag})'}")
            for item in itens:
                linha = {"PROJETO": projeto}
                linha.update(item)
                todos.append(linha)
        _progress(i, total, f"Lendo arquivos... ({i}/{total})")

    tempo_leitura = (datetime.datetime.now() - inicio).total_seconds()
    log(f"\nLeitura concluida em {tempo_leitura:.1f}s "
        f"({total - len(erros)} ok, {len(erros)} com erro)")
    log(f"Total de itens DI encontrados (todos os projetos): {len(todos)}")

    # Tabela 2: deduplicado por CODIGO (1a ocorrencia)
    unicos = []
    vistos = set()
    for linha in todos:
        codigo = linha.get("CODIGO")
        chave = str(codigo).strip() if codigo is not None else None
        if chave is None or chave == "" or chave in vistos:
            continue
        vistos.add(chave)
        unicos.append(linha)

    log(f"Total de itens DI unicos (por CODIGO): {len(unicos)}")

    # --- Monta o workbook de saida ---
    wb_out = openpyxl.Workbook()

    cabecalhos = ["PROJETO"] + COLUNAS_SAIDA

    _escrever_aba(wb_out.active, "Itens DI (Todos)", cabecalhos, todos)
    ws2 = wb_out.create_sheet("Itens DI (Unicos)")
    _escrever_aba(ws2, None, cabecalhos, unicos, ja_existe=True)

    hoje = datetime.date.today().strftime("%y%m%d")
    nome_saida = f"Itens DI - {hoje}.xlsx"
    caminho_saida = os.path.join(pasta_destino, nome_saida)
    wb_out.save(caminho_saida)

    tempo_total = (datetime.datetime.now() - inicio).total_seconds()
    log(f"\nArquivo salvo em: {caminho_saida}")
    log(f"Tempo total: {tempo_total:.1f}s")
    if erros:
        log(f"\n{len(erros)} arquivo(s) com erro (pulados):")
        for nome, msg in erros:
            log(f"  - {nome}: {msg}")
    log("Concluido.")
    _progress(total, total, "Concluido")
    return caminho_saida


def _pool_imap(n_workers, func, arquivos):
    """EN: Thin wrapper so main() doesn't need to manage Pool lifecycle
    inline (keeps the structure close to the Projetos Preliminares
    script)."""
    with mp.Pool(processes=n_workers) as pool:
        for resultado in pool.imap_unordered(func, arquivos):
            yield resultado


def _escrever_aba(ws, titulo_aba, cabecalhos, linhas, ja_existe=False):
    """EN: Writes header + data rows into ws, applies basic formatting
    (header style, borders, autofilter, frozen header, column widths)."""
    if titulo_aba:
        ws.title = titulo_aba

    fonte_header = Font(bold=True, color="404040")
    fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fina = Side(style="thin", color="BFBFBF")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)

    for j, titulo in enumerate(cabecalhos, start=1):
        c = ws.cell(row=1, column=j, value=titulo)
        c.font = fonte_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda

    for i, linha in enumerate(linhas, start=2):
        for j, cab in enumerate(cabecalhos, start=1):
            c = ws.cell(row=i, column=j, value=linha.get(cab))
            c.border = borda
            c.alignment = Alignment(horizontal="center", vertical="center")

    ultima_linha = len(linhas) + 1
    total_colunas = len(cabecalhos)
    ws.auto_filter.ref = f"A1:{get_column_letter(total_colunas)}{ultima_linha}"
    ws.freeze_panes = "A2"

    for col, titulo in enumerate(cabecalhos, start=1):
        largura = max(10, len(str(titulo)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = largura



# =====================================================================
# CRUZAMENTO ITENS DI x PPAP (MONDAY) - logica adaptada de cruzar_ppap.py
# =====================================================================

PASTA_DESTINO_PADRAO = str(Path.home() / "Downloads")

# Colunas do Monday (aba "qualidade", header dinamico - procurado por
# texto exato em qualquer linha das primeiras ~10)
COL_MONDAY_CODIGO = "Name"
COL_MONDAY_FORNECEDOR = "Fornecedor"
COL_MONDAY_DEP = "DEP"
COL_MONDAY_EMBALAGEM = "Embalagem"
COL_MONDAY_STATUS = "Status"

ABA_MONDAY_PREFERIDA = "qualidade"


def _ler_itens_di(caminho_itens_di, log):
    """EN: Reads the "Itens DI (Unicos)" sheet, returns (cabecalhos, linhas)
    where linhas is a list of dicts {cabecalho: valor}."""
    wb = openpyxl.load_workbook(caminho_itens_di, data_only=True, read_only=True)
    try:
        nome_aba = None
        for s in wb.sheetnames:
            if "UNIC" in _norm(s):
                nome_aba = s
                break
        if nome_aba is None:
            nome_aba = wb.sheetnames[0]
            log(f"AVISO: aba 'Itens DI (Unicos)' nao encontrada, usando '{nome_aba}'")

        ws = wb[nome_aba]
        cabecalhos = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        linhas = []
        for row in ws.iter_rows(min_row=2):
            valores = [c.value for c in row]
            if all(v is None for v in valores):
                continue
            linhas.append(dict(zip(cabecalhos, valores)))
        return cabecalhos, linhas
    finally:
        wb.close()


def _ler_monday(caminho_monday, log):
    """EN: Reads the Monday "qualidade" export, locates the header row
    dynamically (looks for COL_MONDAY_CODIGO in column A), returns a dict
    keyed by normalized CODIGO -> {Fornecedor, DEP, Embalagem, Status}."""
    wb = openpyxl.load_workbook(caminho_monday, data_only=True, read_only=False)
    try:
        nome_aba = ABA_MONDAY_PREFERIDA if ABA_MONDAY_PREFERIDA in wb.sheetnames else wb.sheetnames[0]
        ws = wb[nome_aba]

        max_row = ws.max_row or 10
        linha_header = None
        mapa_colunas = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_row, 10)), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if _norm(cell.value) == _norm(COL_MONDAY_CODIGO):
                    linha_header = row_idx
                    break
            if linha_header:
                break

        if linha_header is None:
            log("ERRO: cabecalho 'Name' nao encontrado na planilha Monday.")
            return {}

        for col_idx, cell in enumerate(next(ws.iter_rows(min_row=linha_header, max_row=linha_header)), start=1):
            texto = _norm(cell.value)
            for alvo in (COL_MONDAY_CODIGO, COL_MONDAY_FORNECEDOR, COL_MONDAY_DEP,
                         COL_MONDAY_EMBALAGEM, COL_MONDAY_STATUS):
                if texto == _norm(alvo):
                    mapa_colunas[alvo] = col_idx

        faltando = [c for c in (COL_MONDAY_CODIGO, COL_MONDAY_FORNECEDOR, COL_MONDAY_DEP,
                                 COL_MONDAY_EMBALAGEM, COL_MONDAY_STATUS) if c not in mapa_colunas]
        if faltando:
            log(f"AVISO: colunas nao encontradas no Monday: {', '.join(faltando)}")

        registros = {}
        col_codigo = mapa_colunas.get(COL_MONDAY_CODIGO)
        for linha_dado in range(linha_header + 1, max_row + 1):
            valor_codigo = ws.cell(row=linha_dado, column=col_codigo).value if col_codigo else None
            if valor_codigo is None or str(valor_codigo).strip() == "":
                continue
            chave = str(valor_codigo).strip()
            registros[chave] = {
                "Fornecedor": ws.cell(row=linha_dado, column=mapa_colunas[COL_MONDAY_FORNECEDOR]).value if COL_MONDAY_FORNECEDOR in mapa_colunas else None,
                "DEP": ws.cell(row=linha_dado, column=mapa_colunas[COL_MONDAY_DEP]).value if COL_MONDAY_DEP in mapa_colunas else None,
                "Embalagem": ws.cell(row=linha_dado, column=mapa_colunas[COL_MONDAY_EMBALAGEM]).value if COL_MONDAY_EMBALAGEM in mapa_colunas else None,
                "Status": ws.cell(row=linha_dado, column=mapa_colunas[COL_MONDAY_STATUS]).value if COL_MONDAY_STATUS in mapa_colunas else None,
            }
        return registros
    finally:
        wb.close()


def main_cruzar(caminho_itens_di, caminho_monday, pasta_destino=None, log=print, progress=None):
    pasta_destino = pasta_destino or PASTA_DESTINO_PADRAO

    def _progress(current, total, message=""):
        if progress:
            progress(current, total, message)

    log("=" * 70)
    log("Cruzamento Itens DI x PPAP (Monday) - YYMMDD.xlsx (v1)")
    log("=" * 70)

    if not os.path.isfile(caminho_itens_di):
        log(f"ERRO: arquivo 'Itens DI' nao encontrado: {caminho_itens_di}")
        return None
    if not os.path.isfile(caminho_monday):
        log(f"ERRO: arquivo Monday nao encontrado: {caminho_monday}")
        return None

    inicio = datetime.datetime.now()
    _progress(0, 3, "Lendo Itens DI...")

    cabecalhos_di, itens_di = _ler_itens_di(caminho_itens_di, log)
    log(f"Itens DI (unicos) lidos: {len(itens_di)}")

    _progress(1, 3, "Lendo planilha Monday (Qualidade)...")
    monday = _ler_monday(caminho_monday, log)
    log(f"Registros lidos do Monday: {len(monday)}")

    _progress(2, 3, "Cruzando dados...")

    # Determina a coluna CODIGO no Itens DI (geralmente 2a coluna, apos PROJETO)
    if "CODIGO" in cabecalhos_di:
        col_codigo_nome = "CODIGO"
    else:
        # fallback: procura qualquer cabecalho que normalize para CODIGO
        col_codigo_nome = next((c for c in cabecalhos_di if _norm(c) == "CODIGO"), None)

    if col_codigo_nome is None:
        log("ERRO: coluna 'CODIGO' nao encontrada na planilha de Itens DI.")
        return None

    colunas_monday_saida = ["Fornecedor", "DEP", "Embalagem", "Status"]
    cabecalhos_saida = list(cabecalhos_di) + colunas_monday_saida + ["PPAP_DISPONIVEL"]

    cruzamento = []
    sem_ppap = []
    nao_encontrados = []

    for item in itens_di:
        codigo = item.get(col_codigo_nome)
        chave = str(codigo).strip() if codigo is not None else ""

        linha = dict(item)
        registro_monday = monday.get(chave)

        if registro_monday is None:
            for c in colunas_monday_saida:
                linha[c] = None
            linha["PPAP_DISPONIVEL"] = "NAO"
            cruzamento.append(linha)
            nao_encontrados.append(linha)
            continue

        for c in colunas_monday_saida:
            linha[c] = registro_monday.get(c)

        fornecedor = registro_monday.get("Fornecedor")
        ppap_disponivel = "SIM" if fornecedor and str(fornecedor).strip() else "NAO"
        linha["PPAP_DISPONIVEL"] = ppap_disponivel
        cruzamento.append(linha)

        if ppap_disponivel == "NAO":
            sem_ppap.append(linha)

    com_ppap = sum(1 for l in cruzamento if l["PPAP_DISPONIVEL"] == "SIM")
    log(f"\nCruzamento: {len(cruzamento)} itens")
    log(f"  - Com PPAP disponivel: {com_ppap}")
    log(f"  - Sem PPAP (encontrado no Monday, mas fornecedor vazio): {len(sem_ppap)}")
    log(f"  - Nao encontrados no Monday: {len(nao_encontrados)}")

    # --- Monta workbook de saida ---
    wb_out = openpyxl.Workbook()
    _escrever_aba_ppap(wb_out.active, "Cruzamento", cabecalhos_saida, cruzamento)
    _escrever_aba_ppap(wb_out.create_sheet("Sem PPAP"), None, cabecalhos_saida, sem_ppap)
    _escrever_aba_ppap(wb_out.create_sheet("Nao Encontrados"), None, cabecalhos_saida, nao_encontrados)

    hoje = datetime.date.today().strftime("%y%m%d")
    nome_saida = f"Cruzamento PPAP - {hoje}.xlsx"
    caminho_saida = os.path.join(pasta_destino, nome_saida)
    wb_out.save(caminho_saida)

    tempo_total = (datetime.datetime.now() - inicio).total_seconds()
    log(f"\nArquivo salvo em: {caminho_saida}")
    log(f"Tempo total: {tempo_total:.1f}s")
    log("Concluido.")
    _progress(3, 3, "Concluido")
    return caminho_saida


def _escrever_aba_ppap(ws, titulo_aba, cabecalhos, linhas):
    if titulo_aba:
        ws.title = titulo_aba

    fonte_header = Font(bold=True, color="404040")
    fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fina = Side(style="thin", color="BFBFBF")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)

    fill_sim = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fill_nao = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    col_ppap = None
    for j, titulo in enumerate(cabecalhos, start=1):
        c = ws.cell(row=1, column=j, value=titulo)
        c.font = fonte_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda
        if titulo == "PPAP_DISPONIVEL":
            col_ppap = j

    for i, linha in enumerate(linhas, start=2):
        for j, cab in enumerate(cabecalhos, start=1):
            c = ws.cell(row=i, column=j, value=linha.get(cab))
            c.border = borda
            c.alignment = Alignment(horizontal="center", vertical="center")
            if col_ppap and j == col_ppap:
                c.fill = fill_sim if c.value == "SIM" else fill_nao

    ultima_linha = len(linhas) + 1
    total_colunas = len(cabecalhos)
    ws.auto_filter.ref = f"A1:{get_column_letter(total_colunas)}{ultima_linha}"
    ws.freeze_panes = "A2"

    for col, titulo in enumerate(cabecalhos, start=1):
        largura = max(10, len(str(titulo)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = largura


def executar_gui():
    """
    EN: Unified tkinter GUI with two tabs (ttk.Notebook):
      - "Extrair Itens DI": scans C:\\Projetos, extracts DI items
      - "Cruzar PPAP": cross-references Itens DI (Unicos) with Monday
    Each tab has its own folder/file pickers, progress bar, status line
    and collapsible log, mirroring the original single-purpose GUIs.
    """
    import ttkbootstrap as ttk
    from tkinter import filedialog, messagebox, scrolledtext
    import threading

    root = ttk.Window(themename="cosmo")
    root.title("DNI - Itens DI / Cruzamento PPAP")
    root.geometry("680x340")
    root.resizable(True, True)

    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True, padx=8, pady=(8, 0))

    aba_extrair = ttk.Frame(notebook)
    aba_cruzar = ttk.Frame(notebook)
    notebook.add(aba_extrair, text="Extrair Itens DI")
    notebook.add(aba_cruzar, text="Cruzar PPAP")

    # ================= ABA 1: EXTRAIR ITENS DI =================
    frame_origem = ttk.Frame(aba_extrair)
    frame_origem.pack(fill="x", padx=12, pady=(14, 5))
    ttk.Label(frame_origem, text="Pasta de origem (C:\\Projetos):", width=24, anchor="w").pack(side="left")
    var_origem = ttk.StringVar(value=PASTA_ORIGEM)
    ttk.Entry(frame_origem, textvariable=var_origem).pack(side="left", fill="x", expand=True, padx=5)

    def escolher_origem():
        d = filedialog.askdirectory(initialdir=var_origem.get() or "/")
        if d:
            var_origem.set(d)

    ttk.Button(frame_origem, text="...", command=escolher_origem, width=3, bootstyle="secondary-outline").pack(side="left")

    frame_destino1 = ttk.Frame(aba_extrair)
    frame_destino1.pack(fill="x", padx=12, pady=5)
    ttk.Label(frame_destino1, text="Pasta de destino (Downloads):", width=24, anchor="w").pack(side="left")
    var_destino1 = ttk.StringVar(value=PASTA_DESTINO)
    ttk.Entry(frame_destino1, textvariable=var_destino1).pack(side="left", fill="x", expand=True, padx=5)

    def escolher_destino1():
        d = filedialog.askdirectory(initialdir=var_destino1.get() or "/")
        if d:
            var_destino1.set(d)

    ttk.Button(frame_destino1, text="...", command=escolher_destino1, width=3, bootstyle="secondary-outline").pack(side="left")

    frame_botao1 = ttk.Frame(aba_extrair)
    frame_botao1.pack(fill="x", padx=12, pady=(15, 5))
    btn_extrair = ttk.Button(frame_botao1, text="Extrair Itens DI", width=18, bootstyle="primary")
    btn_extrair.pack(side="left")
    lbl_status1 = ttk.Label(frame_botao1, text="Pronto.", anchor="w")
    lbl_status1.pack(side="left", padx=10, fill="x", expand=True)

    frame_progress1 = ttk.Frame(aba_extrair)
    frame_progress1.pack(fill="x", padx=12, pady=(0, 8))
    pbar1 = ttk.Progressbar(frame_progress1, orient="horizontal", mode="determinate", bootstyle="primary-striped")
    pbar1.pack(fill="x")

    frame_detalhes1 = ttk.Frame(aba_extrair)
    frame_detalhes1.pack(fill="both", expand=True, padx=12, pady=(0, 5))

    log_widget1 = {"w": None}
    detalhes1_visivel = ttk.BooleanVar(value=False)

    def alternar_detalhes1():
        if detalhes1_visivel.get():
            if log_widget1["w"] is None:
                log_widget1["w"] = scrolledtext.ScrolledText(
                    frame_detalhes1, state="disabled", font=("Consolas", 8), height=9
                )
            log_widget1["w"].pack(fill="both", expand=True, pady=(5, 0))
            btn_detalhes1.config(text="▲ Ocultar detalhes")
        else:
            if log_widget1["w"] is not None:
                log_widget1["w"].pack_forget()
            btn_detalhes1.config(text="▼ Mostrar detalhes")

    def toggle_detalhes1():
        detalhes1_visivel.set(not detalhes1_visivel.get())
        alternar_detalhes1()

    btn_detalhes1 = ttk.Button(aba_extrair, text="▼ Mostrar detalhes",
                               bootstyle="link", command=toggle_detalhes1)
    btn_detalhes1.pack(side="bottom", pady=(0, 2))

    def log1(msg):
        def _append():
            if log_widget1["w"] is not None:
                log_widget1["w"].configure(state="normal")
                log_widget1["w"].insert("end", str(msg) + "\n")
                log_widget1["w"].see("end")
                log_widget1["w"].configure(state="disabled")
        root.after(0, _append)

    def progress1(current, total, message=""):
        def _update():
            if total > 0:
                pbar1.configure(maximum=total, value=current)
            if message:
                lbl_status1.config(text=message)
        root.after(0, _update)

    def rodar_extrair():
        btn_extrair.config(state="disabled")
        lbl_status1.config(text="Iniciando...")
        pbar1.configure(value=0)
        if log_widget1["w"] is not None:
            log_widget1["w"].configure(state="normal")
            log_widget1["w"].delete("1.0", "end")
            log_widget1["w"].configure(state="disabled")

        def trabalho():
            try:
                caminho_saida = main(
                    pasta_origem=var_origem.get(),
                    pasta_destino=var_destino1.get(),
                    log=log1,
                    progress=progress1,
                )
            except Exception as e:
                log1(f"\nERRO GERAL: {e}")
                caminho_saida = None

            def finalizar():
                btn_extrair.config(state="normal")
                lbl_status1.config(text="Pronto." if caminho_saida else "Falhou.")
                if caminho_saida:
                    messagebox.showinfo("Concluído", f"Arquivo gerado com sucesso:\n\n{caminho_saida}")
                else:
                    messagebox.showwarning("Atenção",
                        "Processamento finalizado sem gerar arquivo.\n"
                        "Clique em 'Mostrar detalhes' e gere novamente para ver o log.")

            root.after(0, finalizar)

        threading.Thread(target=trabalho, daemon=True).start()

    btn_extrair.config(command=rodar_extrair)

    # ================= ABA 2: CRUZAR PPAP =================
    frame_di = ttk.Frame(aba_cruzar)
    frame_di.pack(fill="x", padx=12, pady=(14, 5))
    ttk.Label(frame_di, text="Arquivo 'Itens DI - ...xlsx':", width=24, anchor="w").pack(side="left")
    var_di = ttk.StringVar(value="")
    ttk.Entry(frame_di, textvariable=var_di).pack(side="left", fill="x", expand=True, padx=5)

    def escolher_di():
        f = filedialog.askopenfilename(
            title="Selecione o arquivo 'Itens DI - ...xlsx'",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls")]
        )
        if f:
            var_di.set(f)

    ttk.Button(frame_di, text="...", command=escolher_di, width=3, bootstyle="secondary-outline").pack(side="left")

    frame_monday = ttk.Frame(aba_cruzar)
    frame_monday.pack(fill="x", padx=12, pady=5)
    ttk.Label(frame_monday, text="Arquivo Monday (Qualidade):", width=24, anchor="w").pack(side="left")
    var_monday = ttk.StringVar(value="")
    ttk.Entry(frame_monday, textvariable=var_monday).pack(side="left", fill="x", expand=True, padx=5)

    def escolher_monday():
        f = filedialog.askopenfilename(
            title="Selecione a planilha exportada do Monday (Qualidade)",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls")]
        )
        if f:
            var_monday.set(f)

    ttk.Button(frame_monday, text="...", command=escolher_monday, width=3, bootstyle="secondary-outline").pack(side="left")

    frame_destino2 = ttk.Frame(aba_cruzar)
    frame_destino2.pack(fill="x", padx=12, pady=5)
    ttk.Label(frame_destino2, text="Pasta de destino (Downloads):", width=24, anchor="w").pack(side="left")
    var_destino2 = ttk.StringVar(value=PASTA_DESTINO_PADRAO)
    ttk.Entry(frame_destino2, textvariable=var_destino2).pack(side="left", fill="x", expand=True, padx=5)

    def escolher_destino2():
        d = filedialog.askdirectory(initialdir=var_destino2.get() or "/")
        if d:
            var_destino2.set(d)

    ttk.Button(frame_destino2, text="...", command=escolher_destino2, width=3, bootstyle="secondary-outline").pack(side="left")

    frame_botao2 = ttk.Frame(aba_cruzar)
    frame_botao2.pack(fill="x", padx=12, pady=(15, 5))
    btn_cruzar = ttk.Button(frame_botao2, text="Cruzar", width=18, bootstyle="info")
    btn_cruzar.pack(side="left")
    lbl_status2 = ttk.Label(frame_botao2, text="Pronto.", anchor="w")
    lbl_status2.pack(side="left", padx=10, fill="x", expand=True)

    frame_progress2 = ttk.Frame(aba_cruzar)
    frame_progress2.pack(fill="x", padx=12, pady=(0, 8))
    pbar2 = ttk.Progressbar(frame_progress2, orient="horizontal", mode="determinate", bootstyle="info-striped")
    pbar2.pack(fill="x")

    frame_detalhes2 = ttk.Frame(aba_cruzar)
    frame_detalhes2.pack(fill="both", expand=True, padx=12, pady=(0, 5))

    log_widget2 = {"w": None}
    detalhes2_visivel = ttk.BooleanVar(value=False)

    def alternar_detalhes2():
        if detalhes2_visivel.get():
            if log_widget2["w"] is None:
                log_widget2["w"] = scrolledtext.ScrolledText(
                    frame_detalhes2, state="disabled", font=("Consolas", 8), height=9
                )
            log_widget2["w"].pack(fill="both", expand=True, pady=(5, 0))
            btn_detalhes2.config(text="▲ Ocultar detalhes")
        else:
            if log_widget2["w"] is not None:
                log_widget2["w"].pack_forget()
            btn_detalhes2.config(text="▼ Mostrar detalhes")

    def toggle_detalhes2():
        detalhes2_visivel.set(not detalhes2_visivel.get())
        alternar_detalhes2()

    btn_detalhes2 = ttk.Button(aba_cruzar, text="▼ Mostrar detalhes",
                               bootstyle="link", command=toggle_detalhes2)
    btn_detalhes2.pack(side="bottom", pady=(0, 2))

    def log2(msg):
        def _append():
            if log_widget2["w"] is not None:
                log_widget2["w"].configure(state="normal")
                log_widget2["w"].insert("end", str(msg) + "\n")
                log_widget2["w"].see("end")
                log_widget2["w"].configure(state="disabled")
        root.after(0, _append)

    def progress2(current, total, message=""):
        def _update():
            if total > 0:
                pbar2.configure(maximum=total, value=current)
            if message:
                lbl_status2.config(text=message)
        root.after(0, _update)

    def rodar_cruzar():
        if not var_di.get() or not var_monday.get():
            messagebox.showwarning("Atenção", "Selecione os dois arquivos (Itens DI e Monday).")
            return

        btn_cruzar.config(state="disabled")
        lbl_status2.config(text="Iniciando...")
        pbar2.configure(value=0)
        if log_widget2["w"] is not None:
            log_widget2["w"].configure(state="normal")
            log_widget2["w"].delete("1.0", "end")
            log_widget2["w"].configure(state="disabled")

        def trabalho():
            try:
                caminho_saida = main_cruzar(
                    caminho_itens_di=var_di.get(),
                    caminho_monday=var_monday.get(),
                    pasta_destino=var_destino2.get(),
                    log=log2,
                    progress=progress2,
                )
            except Exception as e:
                log2(f"\nERRO GERAL: {e}")
                caminho_saida = None

            def finalizar():
                btn_cruzar.config(state="normal")
                lbl_status2.config(text="Pronto." if caminho_saida else "Falhou.")
                if caminho_saida:
                    messagebox.showinfo("Concluído", f"Arquivo gerado com sucesso:\n\n{caminho_saida}")
                else:
                    messagebox.showwarning("Atenção",
                        "Processamento finalizado sem gerar arquivo.\n"
                        "Clique em 'Mostrar detalhes' e gere novamente para ver o log.")

            root.after(0, finalizar)

        threading.Thread(target=trabalho, daemon=True).start()

    btn_cruzar.config(command=rodar_cruzar)

    # --- Rodape: autor (comum as duas abas) ---
    ttk.Label(root, text=f"Autor: {__author__} | DNI - Metalfrio Solutions",
             font=("Segoe UI", 8), bootstyle="secondary").pack(side="bottom", pady=(0, 5))

    root.mainloop()




if __name__ == "__main__":
    mp.freeze_support()
    executar_gui()
