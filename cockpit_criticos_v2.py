"""
Cockpit de Itens Críticos v2 — Metalfrio Solutions
Correções aplicadas:
  - Remove itens com Ruptura = "?" (sem data)
  - Ordena por Ruptura A→Z (cronológico)
  - Qtd Déficit Exato ≠ Qtd Recomendada (cálculos distintos)
  - Formatação centralizada (exceto Descrição: esquerda)
  - % Margem editável na interface — recalcula ao processar
Build: pyinstaller --onefile --windowed cockpit_criticos_v2.py
"""

import os, sys, threading, math
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, date
import pandas as pd, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Identidade Visual ─────────────────────────────────────────────────────────
AZUL="0B1D2F"; VERDE="1CBABE"; BRANCO="FFFFFF"
FAROL_MAP={"Vermelho":("C0392B","FFFFFF"),"Laranja":("E67E22","FFFFFF"),"Amarelo":("F1C40F","1A1A1A"),"Verde":("27AE60","FFFFFF"),"?":("BDC3C7","1A1A1A")}
PRIORIDADE={"Vermelho":0,"Laranja":1,"Amarelo":2,"Verde":3,"?":4}
KEYWORDS_BRANDING=["STELLA","BUD","MICHELOB","CHOPP","BRAHMA","COCA-COLA","PARESA","KIBON","SERIGRAFIA","ADESIVO","CARENAGEM"]
MANUAIS={"Chegada","Qtd. Pedida","Tipo Frete","Fornecedor Alt.","Ação","Carro","Motorista","Status"}
CALC={"Qtd Déficit Exato","Qtd Recomendada","Qtd c/ Margem","% Margem"}

# ── Utilitários ───────────────────────────────────────────────────────────────

def sd(v):
    if v is None or (isinstance(v,float) and pd.isna(v)): return None
    if str(v).strip() in ("","?","nan","—"): return None
    if isinstance(v,(datetime,date)): return v.date() if isinstance(v,datetime) else v
    try: return (datetime(1899,12,30)+pd.Timedelta(days=float(v))).date()
    except: pass
    for f in ("%d/%m/%y","%d/%m/%Y","%Y-%m-%d"):
        try: return datetime.strptime(str(v).strip(),f).date()
        except: pass
    return None

def sf(v,d=0.0):
    try:
        f=float(v)
        return d if(pd.isna(f) or f!=f) else f
    except: return d

def inferir_farol(row):
    f=str(row.get("Farol","")).strip()
    if f in FAROL_MAP: return f
    c=sf(row.get("Cobertura",""),-1)
    if c>=0:
        if c<=0: return "Vermelho"
        if c<=7: return "Laranja"
        if c<=15: return "Amarelo"
        return "Verde"
    return "Vermelho" if sf(row.get("Saldo",""),1)<=0 else "?"

def is_branding(desc): return any(k in str(desc).upper() for k in KEYWORDS_BRANDING)

def pct_dinamico(lt, desc):
    if is_branding(desc): return 0.05
    lt=sf(lt,30)
    if lt<=15: return 0.05
    if lt<=45: return 0.10
    return 0.15

def arred_mul(qtd, mul):
    mul=max(sf(mul,1),1); qtd=sf(qtd,0)
    if qtd<=0: return 0
    return int(math.ceil(qtd/mul)*mul)

def calc_qtd(item, resumo, periodo, pct_override=None):
    res=resumo.get(item,{}); rows=periodo.get(item,[])
    desc=str(res.get("Descrição",""))
    saldos=[sf(r.get("Saldo Projetado",0),9999) for r in rows]
    saldos=[s for s in saldos if s<9999]
    deficit_puro=abs(min(saldos)) if saldos and min(saldos)<0 else 0
    if deficit_puro==0:
        deficit_puro=max(sf(res.get("Reservas Confirmadas",0)),0)
    estq_seg=sf(res.get("Quant Segur",0))
    giro=sf(res.get("Giro HF",0))
    lt=sf(rows[0].get("Lead Time",0)) if rows else sf(res.get("Tmp Segur",30))
    qt_min=max(sf(rows[0].get("Qt-Min",1)) if rows else 1,1)
    qt_mul=max(sf(rows[0].get("Qt-Mul",1)) if rows else 1,1)
    deficit_exato=max(deficit_puro,qt_min)
    base=deficit_puro+estq_seg+(giro*lt)
    qtd_rec=max(arred_mul(base,qt_mul),int(qt_min))
    pct=pct_override if pct_override is not None else pct_dinamico(lt,desc)
    qtd_marg=arred_mul(qtd_rec*(1+pct),qt_mul)
    return int(round(deficit_exato)),qtd_rec,qtd_marg,pct,lt

# ── Leitura ───────────────────────────────────────────────────────────────────

def ler_0604(path):
    xl=pd.ExcelFile(path)
    df=xl.parse(xl.sheet_names[0],skiprows=1,dtype=str)
    df.columns=[str(c).strip() for c in df.columns]
    df=df[df["Item"]!="Item"].dropna(subset=["Item"]).reset_index(drop=True)
    df["Item"]=df["Item"].str.strip()
    return df

def ler_resumo(path):
    xl=pd.ExcelFile(path)
    df=xl.parse(xl.sheet_names[0],dtype=str)
    df.columns=[str(c).strip() for c in df.columns]
    if "Item" in df.columns:
        df=df[df["Item"]!="Item"].dropna(subset=["Item"])
        df["Item"]=df["Item"].str.strip()
        return df.set_index("Item").to_dict(orient="index")
    return {}

def ler_periodo(path):
    xl=pd.ExcelFile(path)
    aba=None
    for nome in xl.sheet_names:
        try:
            cols=list(xl.parse(nome,nrows=3,dtype=str).columns)
            if any("Lead Time" in str(c) for c in cols): aba=nome; break
        except: pass
    if not aba: aba=xl.sheet_names[-1]
    df=xl.parse(aba,dtype=str)
    df.columns=[str(c).strip() for c in df.columns]
    if "Item" not in df.columns:
        for i,row in df.iterrows():
            if any("Lead Time" in str(v) for v in row.values):
                df.columns=df.iloc[i].tolist()
                df=df.iloc[i+1:].reset_index(drop=True); break
    df=df.dropna(subset=["Item"]); df["Item"]=df["Item"].str.strip()
    res={}
    for item,g in df.groupby("Item"): res[item]=g.to_dict(orient="records")
    return res

# ── Consolidação ──────────────────────────────────────────────────────────────

def consolidar(df0604, resumo, periodo, pct_override=None):
    recs=[]
    for _,r04 in df0604.iterrows():
        item=r04["Item"]; res=resumo.get(item,{})
        farol=inferir_farol({"Farol":res.get("Farol",""),"Cobertura":res.get("Cobertura",""),"Saldo":res.get("Saldo","")})
        data_r=sd(r04.get("Data Ruptura"))

        # Remove itens sem data de ruptura definida
        if data_r is None:
            continue

        dias_r=(data_r-date.today()).days
        cob=sf(res.get("Cobertura",""),-1); cob=round(cob,1) if cob>=0 else ""
        custo=sf(res.get("Custo Unitário",""),0); custo=round(custo,4) if 0<custo<999999 else ""
        de,qr,qm,pct,lt=calc_qtd(item,resumo,periodo,pct_override)
        forn=str(res.get("Fornecedores","") or r04.get("Fornecedores","") or "").strip()
        sit="RUPTURA" if farol in ("Vermelho","Laranja") else "OK"

        recs.append({
            "Situação":sit,"Farol":farol,
            "Linha":str(r04.get("Linhas (Onde-se-Usa 30d)","") or "").strip(),
            "Onde Usa":str(r04.get("Produtos (Onde-se-Usa 30d)","") or "").strip(),
            "Autonomia (dias)":cob,
            "Ruptura":data_r,
            "Dias p/ Ruptura":dias_r,
            "Item":item,
            "Descrição":str(r04.get("Descrição","") or res.get("Descrição","") or "").strip(),
            "Fornecedor":forn,
            "Lead Time (dias)":int(lt) if lt else "",
            "Custo Unit. (R$)":custo,
            "Qtd Déficit Exato":de if de else "",
            "Qtd Recomendada":qr if qr else "",
            "Qtd c/ Margem":qm if qm else "",
            "% Margem":f"{int(pct*100)}%",
            "Chegada":"","Qtd. Pedida":"","Tipo Frete":"","Fornecedor Alt.":"",
            "Ação":"","Carro":"","Motorista":"",
            "RESPONSÁVEL":"DNI",
            "Planejador":str(r04.get("Planejador","") or "").strip(),
            "Status":"",
        })

    df=pd.DataFrame(recs)
    # Ordena por Ruptura cronológico (A→Z)
    df["Ruptura"]=pd.to_datetime(df["Ruptura"],errors="coerce")
    df=df.sort_values("Ruptura",ascending=True,na_position="last").reset_index(drop=True)
    df["Ruptura"]=df["Ruptura"].apply(lambda x: x.strftime("%d/%m/%y") if pd.notna(x) else "?")
    return df

# ── Exportação Excel ──────────────────────────────────────────────────────────

def bordas():
    s=Side(style="thin",color="CCCCCC")
    return Border(left=s,right=s,top=s,bottom=s)

def exportar(df, path):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="CRITICOS_DATA"
    ws.sheet_view.showGridLines=False; ws.freeze_panes="A4"
    cols=list(df.columns); last=get_column_letter(len(cols))
    ts=datetime.now().strftime("%d/%m/%Y %H:%M")

    ws.merge_cells(f"A1:{last}1")
    c=ws["A1"]; c.value="GESTÃO DE ITENS CRÍTICOS — METALFRIO SOLUTIONS"
    c.font=Font(name="Segoe UI",bold=True,size=14,color=BRANCO)
    c.fill=PatternFill("solid",fgColor=AZUL)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=30

    ws.merge_cells(f"A2:{last}2")
    c=ws["A2"]; n_vm=len(df[df["Farol"]=="Vermelho"]); n_la=len(df[df["Farol"]=="Laranja"])
    c.value=f"Gerado em {ts}  |  {len(df)} itens  |  Vermelho: {n_vm}  Laranja: {n_la}  |  DNI"
    c.font=Font(name="Segoe UI",size=9,color="8A9BAA")
    c.fill=PatternFill("solid",fgColor=AZUL)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[2].height=15

    for ci,nome in enumerate(cols,1):
        c=ws.cell(row=3,column=ci,value=nome)
        c.font=Font(name="Segoe UI",bold=True,size=9,color=BRANCO)
        bg="1A5276" if nome in MANUAIS else("117A65" if nome in CALC else VERDE)
        c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=bordas()
    ws.row_dimensions[3].height=32

    for i,(_,row) in enumerate(df.iterrows()):
        r=4+i; farol=row.get("Farol","Verde"); sit=row.get("Situação","OK")
        bg_f,fg_f=FAROL_MAP.get(farol,("BDC3C7","1A1A1A"))
        bg_z="F2F4F6" if i%2 else "FFFFFF"

        for ci,col in enumerate(cols,1):
            val=row.get(col,"")
            if isinstance(val,float) and pd.isna(val): val=""
            cell=ws.cell(row=r,column=ci,value=val)
            cell.border=bordas()
            cell.font=Font(name="Segoe UI",size=9)

            # Alinhamento: Descrição = esquerda; todo o resto = centralizado
            if col=="Descrição":
                cell.alignment=Alignment(horizontal="left",vertical="center",indent=1)
            else:
                cell.alignment=Alignment(horizontal="center",vertical="center")

            if col=="Situação":
                cell.fill=PatternFill("solid",fgColor="C0392B" if sit=="RUPTURA" else "27AE60")
                cell.font=Font(name="Segoe UI",size=9,bold=True,color=BRANCO)
            elif col=="Farol":
                cell.fill=PatternFill("solid",fgColor=bg_f)
                cell.font=Font(name="Segoe UI",size=9,bold=True,color=fg_f)
            elif col in MANUAIS:
                cell.fill=PatternFill("solid",fgColor="EBF5FB" if i%2 else "D6EAF8")
            elif col in CALC:
                cell.fill=PatternFill("solid",fgColor="E8F8F5" if i%2 else "D1F2EB")
                cell.font=Font(name="Segoe UI",size=9,bold=True,color="117A65")
            elif col=="Dias p/ Ruptura":
                try:
                    v=int(val)
                    cell.font=Font(name="Segoe UI",size=9,bold=True,
                                   color="C0392B" if v<=7 else("E67E22" if v<=15 else "1A1A1A"))
                except: pass
                cell.fill=PatternFill("solid",fgColor=bg_z)
            else:
                cell.fill=PatternFill("solid",fgColor=bg_z)

        ws.row_dimensions[r].height=16

    # Legenda
    leg=4+len(df)+2
    ws.cell(row=leg,column=1,value="LEGENDA:").font=Font(bold=True,size=9,name="Segoe UI")
    for j,(cor,txt) in enumerate([(VERDE,"Automático (TOTVS)"),("1A5276","Manual"),("117A65","Calculado")],2):
        c=ws.cell(row=leg,column=j,value=txt)
        c.fill=PatternFill("solid",fgColor=cor); c.font=Font(name="Segoe UI",size=9,color=BRANCO)
        c.alignment=Alignment(horizontal="center",vertical="center")

    larg={"Situação":9,"Farol":9,"Linha":22,"Onde Usa":16,"Autonomia (dias)":10,
          "Ruptura":10,"Dias p/ Ruptura":10,"Item":14,"Descrição":46,"Fornecedor":22,
          "Lead Time (dias)":10,"Custo Unit. (R$)":12,
          "Qtd Déficit Exato":14,"Qtd Recomendada":14,"Qtd c/ Margem":14,"% Margem":9,
          "Chegada":12,"Qtd. Pedida":10,"Tipo Frete":11,"Fornecedor Alt.":18,
          "Ação":22,"Carro":12,"Motorista":14,"RESPONSÁVEL":10,"Planejador":24,"Status":34}
    for ci,col in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(ci)].width=larg.get(col,14)

    # Aba KPIs
    wk=wb.create_sheet("📊 KPIs"); wk.sheet_view.showGridLines=False
    wk.merge_cells("A1:B1"); wk["A1"].value="INDICADORES — ITENS CRÍTICOS"
    wk["A1"].font=Font(name="Segoe UI",bold=True,size=12,color=BRANCO)
    wk["A1"].fill=PatternFill("solid",fgColor=AZUL)
    wk["A1"].alignment=Alignment(horizontal="left",indent=1,vertical="center")
    wk.row_dimensions[1].height=26

    def nd(x,lim):
        try: return isinstance(x,(int,float)) and not pd.isna(x) and x<=lim
        except: return False

    kpis=[
        ("Total de Itens Críticos",len(df)),
        ("🔴 Vermelho",len(df[df["Farol"]=="Vermelho"])),
        ("🟠 Laranja",len(df[df["Farol"]=="Laranja"])),
        ("🟡 Amarelo",len(df[df["Farol"]=="Amarelo"])),
        ("🟢 Verde",len(df[df["Farol"]=="Verde"])),
        ("Ruptura já ocorrida",len(df[df["Dias p/ Ruptura"].apply(lambda x: nd(x,-1))])),
        ("Ruptura em ≤ 7 dias",len(df[df["Dias p/ Ruptura"].apply(lambda x: nd(x,7))])),
        ("Ruptura em ≤ 30 dias",len(df[df["Dias p/ Ruptura"].apply(lambda x: nd(x,30))])),
        ("Itens de branding (+5%)",len(df[df["% Margem"]=="5%"])),
        ("Itens margem +10%",len(df[df["% Margem"]=="10%"])),
        ("Itens margem +15%",len(df[df["% Margem"]=="15%"])),
    ]
    for i,(label,val) in enumerate(kpis):
        r=i+2
        ca=wk.cell(row=r,column=1,value=label)
        cb=wk.cell(row=r,column=2,value=val)
        bg="F2F4F6" if i%2 else "FFFFFF"
        for c in (ca,cb):
            c.fill=PatternFill("solid",fgColor=bg); c.border=bordas()
            c.font=Font(name="Segoe UI",size=10)
        cb.font=Font(name="Segoe UI",bold=True,size=11)
        cb.alignment=Alignment(horizontal="center",vertical="center")
        wk.row_dimensions[r].height=20
    wk.column_dimensions["A"].width=38; wk.column_dimensions["B"].width=14

    wb.save(path)

# ── Interface Gráfica ─────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Cockpit de Itens Críticos v2 — Metalfrio Solutions")
        self.geometry("760x600")
        self.resizable(False,False)
        self.configure(bg=f"#{AZUL}")
        self._set_icon()
        self.path_0604=tk.StringVar()
        self.path_0704=tk.StringVar()
        self.path_criticos=tk.StringVar()
        self.path_saida=tk.StringVar(value=os.path.join(
            os.path.expanduser("~"),"Desktop",
            f"CRITICOS_{datetime.now().strftime('%d%m%y')}.xlsx"))
        self.pct_var=tk.StringVar(value="")  # vazio = dinâmico por LT
        self.status=tk.StringVar(value="Pronto.")
        self.progresso=tk.DoubleVar(value=0)
        self._build()

    def _set_icon(self):
        try:
            base=getattr(sys,"_MEIPASS",os.path.dirname(__file__))
            ico=os.path.join(base,"icon_dni.ico")
            if os.path.exists(ico): self.iconbitmap(ico)
        except: pass

    def _build(self):
        tk.Frame(self,bg=f"#{AZUL}",height=4).pack(fill="x")
        hdr=tk.Frame(self,bg=f"#{AZUL}",pady=12); hdr.pack(fill="x")
        tk.Label(hdr,text="COCKPIT DE ITENS CRÍTICOS",
                 font=("Segoe UI",17,"bold"),bg=f"#{AZUL}",fg=f"#{VERDE}").pack()
        tk.Label(hdr,text="Metalfrio Solutions  •  DNI — Desenvolvimento de Novos Itens",
                 font=("Segoe UI",9),bg=f"#{AZUL}",fg="#5A8FAA").pack(pady=(2,0))
        tk.Frame(self,bg=f"#{VERDE}",height=3).pack(fill="x")

        body=tk.Frame(self,bg="#EFF2F5",padx=30,pady=16); body.pack(fill="both",expand=True)
        body.columnconfigure(0,weight=1)

        self._campo(body,"1.  ESPL0604  —  Listagem Itens Críticos",self.path_0604,0)
        self._campo(body,"2.  ESPL0704  —  Demonstrativo Cálculo MRP",self.path_0704,2)
        self._campo(body,"3.  CRITICOS existente  —  opcional (merge de campos manuais)",
                    self.path_criticos,4,opcional=True)
        self._campo_saida(body,6)

        # Campo % Margem editável
        pct_frame=tk.Frame(body,bg="#EFF2F5"); pct_frame.grid(row=8,column=0,columnspan=3,sticky="w",pady=(10,0))
        tk.Label(pct_frame,text="% Margem sobre Qtd Recomendada:",
                 font=("Segoe UI",9,"bold"),bg="#EFF2F5",fg=f"#{AZUL}").pack(side="left")
        tk.Entry(pct_frame,textvariable=self.pct_var,width=6,
                 font=("Segoe UI",10),relief="solid",bd=1,justify="center").pack(side="left",padx=(8,4))
        tk.Label(pct_frame,text="%",font=("Segoe UI",9),bg="#EFF2F5",fg=f"#{AZUL}").pack(side="left")
        tk.Label(pct_frame,
                 text="  (deixe vazio para % dinâmico por Lead Time: ≤15d=5% | 16–45d=10% | >45d=15%)",
                 font=("Segoe UI",8),bg="#EFF2F5",fg="#888").pack(side="left")

        # Legenda
        leg=tk.Frame(body,bg="#EFF2F5"); leg.grid(row=9,column=0,columnspan=3,pady=(8,0),sticky="w")
        for cor_hex,texto in [(VERDE,"Automático (TOTVS)"),("1A5276","Campo manual"),("117A65","Calculado pelo sistema")]:
            f=tk.Frame(leg,bg=f"#{cor_hex}",width=12,height=12); f.pack(side="left",padx=(0,3))
            tk.Label(leg,text=texto,font=("Segoe UI",8),bg="#EFF2F5",fg="#555").pack(side="left",padx=(0,12))

        btn_frame=tk.Frame(body,bg="#EFF2F5"); btn_frame.grid(row=10,column=0,columnspan=3,pady=12)
        self.btn=tk.Button(btn_frame,text="▶   PROCESSAR E EXPORTAR",
                           font=("Segoe UI",11,"bold"),bg=f"#{VERDE}",fg="white",
                           activebackground="#139A9D",activeforeground="white",
                           relief="flat",padx=28,pady=10,cursor="hand2",command=self._iniciar)
        self.btn.pack()

        ttk.Progressbar(body,variable=self.progresso,maximum=100,mode="determinate",length=680).grid(
            row=11,column=0,columnspan=3,pady=(0,5))
        tk.Label(body,textvariable=self.status,font=("Segoe UI",9),bg="#EFF2F5",fg="#555").grid(
            row=12,column=0,columnspan=3)

        tk.Frame(self,bg=f"#{AZUL}",height=3).pack(fill="x",side="bottom")
        rod=tk.Frame(self,bg=f"#{AZUL}",pady=5); rod.pack(fill="x",side="bottom")
        tk.Label(rod,text="DNI  •  Metalfrio Solutions  •  v2.1",
                 font=("Segoe UI",8),bg=f"#{AZUL}",fg="#3D6880").pack()

    def _campo(self,parent,label,var,row,opcional=False):
        lbl=label+("  ⬡ opcional" if opcional else "")
        tk.Label(parent,text=lbl,font=("Segoe UI",9,"bold"),bg="#EFF2F5",fg=f"#{AZUL}").grid(
            row=row,column=0,columnspan=3,sticky="w",pady=(8,2))
        tk.Entry(parent,textvariable=var,width=76,font=("Segoe UI",9),relief="solid",bd=1).grid(
            row=row+1,column=0,columnspan=2,sticky="ew")
        tk.Button(parent,text="📂 Selecionar",font=("Segoe UI",9),bg=f"#{AZUL}",fg="white",
                  relief="flat",padx=10,cursor="hand2",
                  command=lambda v=var:self._abrir(v)).grid(row=row+1,column=2,padx=(6,0))

    def _campo_saida(self,parent,row):
        tk.Label(parent,text="4.  Arquivo de saída  (.xlsx)",font=("Segoe UI",9,"bold"),
                 bg="#EFF2F5",fg=f"#{AZUL}").grid(row=row,column=0,columnspan=3,sticky="w",pady=(8,2))
        tk.Entry(parent,textvariable=self.path_saida,width=76,
                 font=("Segoe UI",9),relief="solid",bd=1).grid(row=row+1,column=0,columnspan=2,sticky="ew")
        tk.Button(parent,text="💾 Salvar como",font=("Segoe UI",9),bg=f"#{AZUL}",fg="white",
                  relief="flat",padx=10,cursor="hand2",command=self._salvar_como).grid(
            row=row+1,column=2,padx=(6,0))

    def _abrir(self,var):
        f=filedialog.askopenfilename(title="Selecionar arquivo",
                                     filetypes=[("Excel","*.xlsx *.xls"),("Todos","*.*")])
        if f: var.set(f)

    def _salvar_como(self):
        f=filedialog.asksaveasfilename(title="Salvar como",defaultextension=".xlsx",
                                        filetypes=[("Excel","*.xlsx")])
        if f: self.path_saida.set(f)

    def _iniciar(self):
        erros=[]
        if not self.path_0604.get() or not os.path.exists(self.path_0604.get()):
            erros.append("• Arquivo ESPL0604 não encontrado.")
        if not self.path_0704.get() or not os.path.exists(self.path_0704.get()):
            erros.append("• Arquivo ESPL0704 não encontrado.")
        if not self.path_saida.get():
            erros.append("• Defina o arquivo de saída.")
        # Valida % se preenchido
        pct_txt=self.pct_var.get().strip().replace("%","")
        pct_override=None
        if pct_txt:
            try:
                pct_override=float(pct_txt)/100
                if not 0<pct_override<1:
                    erros.append("• % Margem deve estar entre 1% e 99%.")
            except ValueError:
                erros.append("• % Margem inválido — use apenas números (ex: 10).")
        if erros:
            messagebox.showerror("Atenção","\n".join(erros)); return
        self._pct_override=pct_override
        self.btn.config(state="disabled")
        threading.Thread(target=self._processar,daemon=True).start()

    def _up(self,pct,msg):
        self.progresso.set(pct); self.status.set(msg); self.update_idletasks()

    def _processar(self):
        try:
            self._up(8,"Lendo ESPL0604…")
            df0604=ler_0604(self.path_0604.get())
            self._up(25,"Lendo ESPL0704 — resumo…")
            resumo=ler_resumo(self.path_0704.get())
            self._up(42,"Lendo ESPL0704 — Item × Período…")
            periodo=ler_periodo(self.path_0704.get())
            self._up(60,"Consolidando e calculando quantidades…")
            df=consolidar(df0604,resumo,periodo,self._pct_override)
            self._up(80,"Exportando Excel formatado…")
            exportar(df,self.path_saida.get())
            n_vm=len(df[df["Farol"]=="Vermelho"]); n_la=len(df[df["Farol"]=="Laranja"])
            pct_info=f"{int(self._pct_override*100)}% fixo" if self._pct_override else "dinâmico por LT"
            self._up(100,f"✅  {len(df)} itens | 🔴 {n_vm} | 🟠 {n_la} | % {pct_info}")
            messagebox.showinfo("Concluído",
                f"CRITICOS_DATA gerado!\n\n"
                f"Itens processados:  {len(df)}\n"
                f"🔴 Vermelho:  {n_vm}\n"
                f"🟠 Laranja:   {n_la}\n"
                f"🟡 Amarelo:   {len(df[df['Farol']=='Amarelo'])}\n"
                f"🟢 Verde:     {len(df[df['Farol']=='Verde'])}\n\n"
                f"% Margem aplicado:  {pct_info}\n"
                f"Arquivo: {self.path_saida.get()}")
            try: os.startfile(os.path.dirname(self.path_saida.get()))
            except: pass
        except Exception as e:
            self._up(0,f"❌  Erro: {e}")
            messagebox.showerror("Erro",str(e))
        finally:
            self.btn.config(state="normal")

if __name__=="__main__":
    App().mainloop()
